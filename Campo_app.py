import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import date
import os
import io
import zipfile
import folium
from streamlit_folium import st_folium

st.set_page_config(page_title="SUPERVISIÓN CANAL TRADICIONAL MAU VERSION", layout="wide", page_icon=None)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600&family=DM+Mono&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #f5f4f0; }
    .block-container { padding: 2rem 2rem 2rem 2rem; }
    h1 { font-size: 1.6rem !important; font-weight: 600 !important; color: #1a1a1a !important; }
    h2 { font-size: 1.1rem !important; font-weight: 500 !important; color: #333 !important; }
    .kpi-box { background: white; border-radius: 12px; padding: 1.2rem 1.4rem; border: 1px solid #e8e6e0; }
    .kpi-label { font-size: 13px; text-transform: uppercase; letter-spacing: .08em; color: #888; margin-bottom: 4px; }
    .kpi-value { font-size: 32px; font-weight: 600; color: #1a1a1a; }
    .kpi-sub { font-size: 13px; color: #aaa; margin-top: 2px; }
    .stButton > button { background: #1a1a1a !important; color: white !important; border: none !important;
        border-radius: 8px !important; font-family: 'DM Sans', sans-serif !important;
        font-weight: 500 !important; padding: .6rem 2rem !important; font-size: 15px !important; }
    .stButton > button:hover { background: #333 !important; }
    div[data-testid="stCheckbox"] label { font-size: 13px !important; }
    .leyenda-box { background: #f0f0eb; border-radius: 8px; padding: .5rem 1rem; font-size: 13px; color: #666; margin-bottom: 12px; }
    .btn-danger > button { background: #e05252 !important; color: white !important; border: none !important; }
    .btn-danger > button:hover { background: #c0392b !important; }
    .edit-info { background: #e8f4fd; border: 1px solid #b3d9f5; border-radius: 8px; padding: 8px 14px;
        font-size: 13px; color: #1a5276; margin-bottom: 8px; }
    .filter-info { background: #fef9e7; border: 1px solid #f9e79f; border-radius: 8px; padding: 8px 14px;
        font-size: 13px; color: #7d6608; margin-bottom: 8px; }
</style>
""", unsafe_allow_html=True)

CSV_FILE = "visitas.csv"
IMG_FOLDER = "imagenes_visita"
os.makedirs(IMG_FOLDER, exist_ok=True)

COLUMNAS = [
    "Fecha", "Codigo_PDC", "Nombre_Cliente", "Giro_Negocio",
    "Vendedor", "Codigo_Vendedor", "Mesa", "Zona", "Latitud", "Longitud",
    "OREO_34GR", "OREO_54GR", "OREO_ROLLO", "RITZ_ROLLO", "RITZ_PACK",
    "FIELD_CC", "FIELD_DP", "FIELD_VAIN", "CLUB_SOCIAL_TRA",
    "OREO_FRESA_PACK", "OREO_FRESA_ROLLO",
    "OREO_CHOCO_LIMON_PACK", "OREO_CHOCO_LIMON_ROLLO",
    "CLUB_SOCIAL_SAB", "ROLLO_GOLDEN", "ROLLO_CHOCOLATE",
    "TRIDENT_5s", "TRIDENT_EVUP", "HALLS_12s", "HALLS_100s", "CHICLETS_2S", "BUBBALOO",
    "LEGOS_GC", "TOBOGAN_RITZ_OREO", "EXHIB_KIWI", "RITRAZ", "MEGA_KIWI",
    "EXHIBIDOR_OTROS", "EXHIBIDOR_OTROS_DESC",
    "CONT_LEGOS_GC", "CONT_TOBOGAN_RITZ_OREO", "CONT_EXHIB_KIWI", "Causa_Contaminacion",
    "Visibilidad_Legos", "Visibilidad_Tobogan", "Visibilidad_Kiwi", "Visibilidad_Otros",
    "Visibilidad_Otros_Desc",
    "Colocacion_Terceros", "Marca_Tercero",
    "Efectividad_Soles", "Tiempo_PDC", "Imagen_Path",
]

# Mapeo de giros a nombres legibles para el eje X
GIRO_LABELS = {
    "1 - Bodega": "Bodega",
    "2 - Minimarket / Tiendas": "Minimarket",
    "3 - Kiosko": "Kiosko",
    "4 - Especializados (Panificadora, Horeca, Internet...)": "Especializados",
    "5 - Otros (Puesto de mercado, Centros Educativos...)": "Otros",
}

def giro_short(giro_str):
    """Retorna el nombre corto del giro, compatible con label original o limpio."""
    if pd.isna(giro_str):
        return "Sin Giro"
    s = str(giro_str).strip()
    for k, v in GIRO_LABELS.items():
        if k.lower() in s.lower() or s.lower() in k.lower():
            return v
    # fallback: quitar número inicial
    import re
    return re.sub(r"^\d+ - ", "", s).strip()


def cargar_datos():
    if os.path.exists(CSV_FILE):
        return pd.read_csv(CSV_FILE)
    return pd.DataFrame(columns=COLUMNAS)

def guardar_registro(registro):
    df = cargar_datos()
    nuevo = pd.DataFrame([registro])
    df = pd.concat([df, nuevo], ignore_index=True)
    df.to_csv(CSV_FILE, index=False)

def eliminar_historial():
    if os.path.exists(CSV_FILE):
        os.remove(CSV_FILE)
    if os.path.exists(IMG_FOLDER):
        for f in os.listdir(IMG_FOLDER):
            os.remove(os.path.join(IMG_FOLDER, f))

for _k, _v in {
    "pagina": "formulario", "confirmar_eliminar": False, "gps_lat": "", "gps_lon": "",
    "snapshots": {}, "geo_resultados": [], "buscar_trigger": False,
    "df_editado": None,           # DataFrame editable
    "exhibidor_seleccionado": None,  # Filtro activo en gráfico exhibidor→giro
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

SNAPSHOTS_FILE = "snapshots.pkl"

def guardar_snapshots_disco():
    import pickle
    with open(SNAPSHOTS_FILE, "wb") as f:
        pickle.dump(st.session_state.snapshots, f)

def cargar_snapshots_disco():
    import pickle
    if os.path.exists(SNAPSHOTS_FILE):
        with open(SNAPSHOTS_FILE, "rb") as f:
            return pickle.load(f)
    return {}

if not st.session_state.snapshots and os.path.exists(SNAPSHOTS_FILE):
    st.session_state.snapshots = cargar_snapshots_disco()


# ═══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL: GENERAR DASHBOARD EXCEL CON GRÁFICOS
# ═══════════════════════════════════════════════════════════════
def generar_dashboard_excel(df_f, ticket_calc, fecha_desde, fecha_hasta, filtro_vendedor,
                             total_visitas, total_ventas, ticket_prom_global,
                             tiempo_prom, pct_con_terceros, pct_concreto, data_exhib):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from collections import Counter

    PALETTE = ["#4472C4", "#e05252", "#6a9e4f", "#FF7F0E", "#7b5ea7",
               "#FFC107", "#00BCD4", "#FF69B4", "#8BC34A", "#FF5722",
               "#9C27B0", "#03A9F4", "#CDDC39", "#795548", "#607D8B"]

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    dash_buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Operativo"

    def insert_image_bytes(img_bytes, anchor_cell, width_px=480, height_px=320):
        img_io = io.BytesIO(img_bytes)
        img    = XLImage(img_io)
        img.width  = width_px
        img.height = height_px
        ws.add_image(img, anchor_cell)

    def fig_to_bytes(fig):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor="white")
        buf.seek(0)
        return buf.read()

    def section_header(row, text, color="4472C4"):
        ws.merge_cells(f"A{row}:L{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        return row + 1

    def reserve_rows(start_row, n, height=15):
        for r in range(start_row, start_row + n):
            ws.row_dimensions[r].height = height
        return start_row + n

    IMG_FULL_W = 980
    IMG_HALF_W = 480
    IMG_H      = 330
    IMG_H_MAP  = 430
    IMG_H_WIDE = 330

    def chart_style(ax, title, ylabel="", xlabel=""):
        ax.set_title(title, fontsize=13, fontweight="bold", pad=10)
        if ylabel: ax.set_ylabel(ylabel, fontsize=10)
        if xlabel: ax.set_xlabel(xlabel, fontsize=10)
        ax.yaxis.grid(True, linestyle="--", alpha=0.45, zorder=0)
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    # CABECERA
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="DASHBOARD OPERATIVO — SUPERVISIÓN CANAL TRADICIONAL")
    c.font      = Font(bold=True, size=18, color="FFFFFF", name="Arial")
    c.fill      = PatternFill("solid", fgColor="1a1a1a")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1,
                 value=f"Período: {fecha_desde}  →  {fecha_hasta}  |  Vendedor: {filtro_vendedor}")
    c2.font      = Font(size=12, color="555555", name="Arial")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:L3")
    c3 = ws.cell(row=3, column=1, value="INDICADORES CLAVE")
    c3.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    c3.fill      = PatternFill("solid", fgColor="4472C4")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 26

    kpi_data = [
        ("EFECTIVIDAD TOTAL (S/)", f"S/ {total_ventas:,.2f}",      "7b5ea7"),
        ("TICKET PROMEDIO (S/)",   f"S/ {ticket_prom_global:.2f}", "4472C4"),
        ("% CONCRETO VENTA",       f"{pct_concreto:.0f}%",         "6a9e4f"),
        ("TOTAL VISITAS",          str(total_visitas),             "FF7F0E"),
        ("TIEMPO PROM. PDC",       f"{tiempo_prom:.0f} min",       "e05252"),
        ("% CON TERCEROS",         f"{pct_con_terceros:.0f}%",     "FFC107"),
    ]
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 42
    ws.row_dimensions[6].height = 6

    for ci, (lbl, val, color) in enumerate(kpi_data, start=1):
        cs = (ci - 1) * 2 + 1
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=cs + 1)
        cl = ws.cell(row=4, column=cs, value=lbl)
        cl.font      = Font(bold=True, size=10, color="FFFFFF", name="Arial")
        cl.fill      = PatternFill("solid", fgColor=color)
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=cs + 1)
        cv = ws.cell(row=5, column=cs, value=val)
        cv.font      = Font(bold=True, size=20, color=color, name="Arial")
        cv.alignment = Alignment(horizontal="center", vertical="center")
        cv.fill      = PatternFill("solid", fgColor="F8F8F8")

    # TICKET TABLE
    row_cur = 7
    ws.merge_cells(f"A{row_cur}:L{row_cur}")
    ctk = ws.cell(row=row_cur, column=1, value="TICKET PROMEDIO POR VENDEDOR Y DÍA")
    ctk.font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ctk.fill      = PatternFill("solid", fgColor="4472C4")
    ctk.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_cur].height = 24

    for ci, h in enumerate(["Vendedor", "Fecha", "Ventas del Día (S/)",
                             "Clientes Visitados", "Ticket Promedio (S/)"], 1):
        c = ws.cell(row=row_cur + 1, column=ci, value=h)
        c.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        c.fill      = PatternFill("solid", fgColor="6a9e4f")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        ws.row_dimensions[row_cur + 1].height = 22

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    for ri, row_d in ticket_calc.iterrows():
        r = row_cur + 2 + ri
        for ci2, val in enumerate([row_d["Vendedor"], row_d["Fecha_str"],
                                    round(row_d["Ventas_Dia"], 2),
                                    int(row_d["Clientes_Dia"]),
                                    round(row_d["Ticket_Calculado"], 2)], 1):
            c = ws.cell(row=r, column=ci2, value=val)
            c.border    = border
            c.font      = Font(size=10, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 1:
                c.fill = alt_fill
        ws.row_dimensions[r].height = 18

    row_cur = row_cur + 2 + len(ticket_calc) + 1

    # MAPA
    if "Latitud" in df_f.columns and "Longitud" in df_f.columns:
        df_map = df_f.copy()
        df_map["Latitud"]  = pd.to_numeric(df_map["Latitud"],  errors="coerce")
        df_map["Longitud"] = pd.to_numeric(df_map["Longitud"], errors="coerce")
        df_map = df_map.dropna(subset=["Latitud", "Longitud"])
        if not df_map.empty:
            row_cur = section_header(row_cur, "📍 MAPA DE VISITAS — DISTRIBUCIÓN GEOGRÁFICA", "4472C4")
            cs_map  = row_cur
            row_cur = reserve_rows(row_cur, 30, 15)
            zonas_u = df_map["Zona"].dropna().unique().tolist() if "Zona" in df_map.columns else ["Sin Zona"]
            zc_map  = {z: PALETTE[i % len(PALETTE)] for i, z in enumerate(zonas_u)}
            fig_m, ax_m = plt.subplots(figsize=(14, 6))
            for zona in zonas_u:
                grp = df_map[df_map["Zona"] == zona] if "Zona" in df_map.columns else df_map
                ax_m.scatter(grp["Longitud"], grp["Latitud"],
                             c=zc_map.get(zona, "#888"), s=100, label=zona,
                             alpha=0.88, edgecolors="white", linewidths=0.8, zorder=5)
            for _, rm in df_map.iterrows():
                ax_m.annotate(str(rm.get("Nombre_Cliente", ""))[:20],
                              (rm["Longitud"], rm["Latitud"]),
                              textcoords="offset points", xytext=(6, 4),
                              fontsize=7, alpha=0.8)
            chart_style(ax_m, "MAPA DE VISITAS — DISTRIBUCIÓN GEOGRÁFICA",
                        ylabel="Latitud", xlabel="Longitud")
            ax_m.legend(title="Zona", fontsize=9, title_fontsize=10,
                        loc="upper right", framealpha=0.9, edgecolor="#ccc")
            ax_m.grid(True, linestyle="--", alpha=0.35)
            plt.tight_layout()
            insert_image_bytes(fig_to_bytes(fig_m), f"A{cs_map}", IMG_FULL_W, IMG_H_MAP)
            plt.close(fig_m)
            row_cur += 1

    # GIRO + EFECTIVIDAD
    row_cur = section_header(row_cur, "GIRO DE NEGOCIO  |  EFECTIVIDAD DE VISITAS", "4472C4")
    cs_ge   = row_cur
    row_cur = reserve_rows(row_cur, 24, 15)

    if "Giro_Negocio" in df_f.columns:
        df_giro = df_f["Giro_Negocio"].value_counts().reset_index()
        df_giro.columns = ["Giro", "Visitas"]
        df_giro["Giro_Short"] = df_giro["Giro"].apply(giro_short)
        total_g = df_giro["Visitas"].sum()
        fig_gi, ax_gi = plt.subplots(figsize=(8, 5.5))
        wc = [PALETTE[i % len(PALETTE)] for i in range(len(df_giro))]
        wedges, _, autos = ax_gi.pie(
            df_giro["Visitas"], labels=None,
            autopct=lambda p: f"{p:.0f}%\n({int(round(p * total_g / 100))})",
            colors=wc, startangle=90, pctdistance=0.72,
            textprops={"fontsize": 9},
        )
        for at in autos:
            at.set_fontweight("bold")
        ax_gi.legend(wedges,
                     [f"{g}  ({v})" for g, v in zip(df_giro["Giro_Short"], df_giro["Visitas"])],
                     loc="center left", bbox_to_anchor=(1, 0.5), fontsize=8)
        ax_gi.set_title("GIROS DE NEGOCIO", fontsize=13, fontweight="bold", pad=10)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_gi), f"A{cs_ge}", IMG_HALF_W, IMG_H)
        plt.close(fig_gi)

    if "Concreto" in df_f.columns:
        df_ef  = df_f["Concreto"].value_counts().reset_index()
        df_ef.columns = ["Estado", "Cantidad"]
        total_ef = df_ef["Cantidad"].sum()
        fig_ef, ax_ef = plt.subplots(figsize=(8, 5.5))
        col_ef = ["#7b5ea7" if e == "CONCRETO" else "#e05252" for e in df_ef["Estado"]]
        bars_ef = ax_ef.barh(df_ef["Estado"], df_ef["Cantidad"],
                             color=col_ef, edgecolor="white", height=0.55)
        max_ef = df_ef["Cantidad"].max() if not df_ef.empty else 1
        for bar, cnt in zip(bars_ef, df_ef["Cantidad"]):
            pct = round(cnt / total_ef * 100) if total_ef > 0 else 0
            ax_ef.text(bar.get_width() + max_ef * 0.03,
                       bar.get_y() + bar.get_height() / 2,
                       f"{cnt}\n({pct}%)", va="center", ha="left",
                       fontsize=11, fontweight="bold")
        chart_style(ax_ef, "EFECTIVIDAD DE VISITAS", xlabel="Cantidad de Visitas")
        ax_ef.set_xlim(0, max_ef * 1.45)
        ax_ef.xaxis.grid(True, linestyle="--", alpha=0.45)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_ef), f"G{cs_ge}", IMG_HALF_W, IMG_H)
        plt.close(fig_ef)

    row_cur += 1

    # EXHIBIDORES
    exhib_cols_dict = {
        "LEGOS_GC": "LEGOS G&C", "TOBOGAN_RITZ_OREO": "TOBOGÁN Ritz/Oreo",
        "EXHIB_KIWI": "EXHIB KIWI", "RITRAZ": "RITRAZ",
        "MEGA_KIWI": "MEGA KIWI", "EXHIBIDOR_OTROS": "OTROS",
    }

    row_cur = section_header(row_cur, "COLOCACIÓN DE EXHIBIDORES  |  TIPO DE NEGOCIO POR EXHIBIDOR", "FF7F0E")
    cs_ex   = row_cur
    row_cur = reserve_rows(row_cur, 24, 15)

    labels_e = [d["Exhibidor"] for d in data_exhib]
    vals_e   = [d["Cantidad"]  for d in data_exhib]
    pcts_e   = [d["Pct"]       for d in data_exhib]
    max_e    = max(vals_e) if vals_e else 1

    fig_ex, ax_ex = plt.subplots(figsize=(10, 5))
    bars_ex = ax_ex.bar(labels_e, vals_e,
                        color=[PALETTE[i % len(PALETTE)] for i in range(len(labels_e))],
                        edgecolor="white", linewidth=1.2, zorder=3)
    for bar, cnt, pct in zip(bars_ex, vals_e, pcts_e):
        ax_ex.text(bar.get_x() + bar.get_width() / 2,
                   bar.get_height() + max_e * 0.025,
                   f"{cnt}\n({pct}%)",
                   ha="center", va="bottom", fontsize=10, fontweight="bold")
    chart_style(ax_ex, "COLOCACIÓN DE EXHIBIDORES", ylabel="Cantidad de Visitas")
    ax_ex.set_ylim(0, max_e * 1.4)
    plt.xticks(rotation=22, ha="right", fontsize=9)
    plt.tight_layout()
    insert_image_bytes(fig_to_bytes(fig_ex), f"A{cs_ex}", IMG_HALF_W, IMG_H)
    plt.close(fig_ex)

    # Tipo de Negocio por Exhibidor
    import numpy as np
    exhib_giro_data = []
    for col_ex2, label_ex2 in exhib_cols_dict.items():
        if col_ex2 in df_f.columns and "Giro_Negocio" in df_f.columns:
            df_ex_s = df_f[pd.to_numeric(df_f[col_ex2], errors="coerce").fillna(0) > 0].copy()
            if df_ex_s.empty:
                continue
            df_ex_s["Giro_Short"] = df_ex_s["Giro_Negocio"].apply(giro_short)
            for giro, grp in df_ex_s.groupby("Giro_Short"):
                exhib_giro_data.append({"Exhibidor": label_ex2, "Giro": giro, "Cantidad": len(grp)})

    if exhib_giro_data:
        df_eg       = pd.DataFrame(exhib_giro_data)
        giros_u     = sorted(df_eg["Giro"].unique().tolist())
        exhibs_u    = list(exhib_cols_dict.values())
        exhibs_u    = [e for e in exhibs_u if e in df_eg["Exhibidor"].unique()]
        n_giros     = len(giros_u)
        n_exhibs    = len(exhibs_u)
        color_exhib = {e: PALETTE[i % len(PALETTE)] for i, e in enumerate(exhibs_u)}
        x_pos  = np.arange(n_giros)
        width  = 0.8 / max(n_exhibs, 1)

        fig_eg2, ax_eg2 = plt.subplots(figsize=(max(10, n_giros * 2.2), 5))
        for ei, exhib in enumerate(exhibs_u):
            vals_eg = []
            for giro in giros_u:
                sub = df_eg[(df_eg["Exhibidor"] == exhib) & (df_eg["Giro"] == giro)]
                vals_eg.append(int(sub["Cantidad"].sum()) if not sub.empty else 0)
            offset  = (ei - n_exhibs / 2 + 0.5) * width
            bars_eg = ax_eg2.bar(x_pos + offset, vals_eg, width * 0.92,
                                 label=exhib, color=color_exhib[exhib],
                                 edgecolor="white", linewidth=0.8, zorder=3)
            max_eg_val = max(vals_eg) if vals_eg else 1
            for bar, cnt in zip(bars_eg, vals_eg):
                if cnt > 0:
                    pct_eg = round(cnt / total_visitas * 100, 1)
                    ax_eg2.text(bar.get_x() + bar.get_width() / 2,
                                bar.get_height() + max_eg_val * 0.04,
                                f"{cnt}\n({pct_eg}%)",
                                ha="center", va="bottom", fontsize=8, fontweight="bold")

        ax_eg2.set_xticks(x_pos)
        ax_eg2.set_xticklabels(giros_u, fontsize=9, rotation=15, ha="right")
        chart_style(ax_eg2, "COLOCACIÓN DE EXHIBIDORES POR GIRO DE NEGOCIO",
                    ylabel="Cantidad de Visitas")
        all_vals_eg = [row_d["Cantidad"] for row_d in exhib_giro_data]
        ax_eg2.set_ylim(0, (max(all_vals_eg) if all_vals_eg else 1) * 1.5)
        ax_eg2.legend(title="Exhibidor", fontsize=8, title_fontsize=9, loc="upper right", ncol=2)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_eg2), f"G{cs_ex}", IMG_HALF_W, IMG_H)
        plt.close(fig_eg2)

    row_cur += 1

    # TERCEROS + MARCAS
    row_cur = section_header(row_cur, "COLOCACIÓN DE TERCEROS  |  MARCAS DE COMPETENCIA", "7b5ea7")
    cs_tc   = row_cur
    row_cur = reserve_rows(row_cur, 24, 15)

    if "Colocacion_Terceros" in df_f.columns and "Giro_Negocio" in df_f.columns:
        df_tg = df_f.copy()
        df_tg["Giro_Short"] = df_tg["Giro_Negocio"].apply(giro_short).str.upper()
        df_con = (df_tg[df_tg["Colocacion_Terceros"] == "Sí"]
                  .groupby("Giro_Short").size().reset_index(name="N"))
        sin_n    = int((df_tg["Colocacion_Terceros"] != "Sí").sum())
        lbl_pie  = df_con["Giro_Short"].tolist() + ["SIN COLOCACION"]
        val_pie  = df_con["N"].tolist() + [sin_n]
        tot_pie  = sum(val_pie)
        col_pie  = [PALETTE[i % len(PALETTE)] for i in range(len(df_con))] + ["#4CAF50"]

        fig_tc, ax_tc = plt.subplots(figsize=(8, 5.5))
        wedges_t, _, autos_t = ax_tc.pie(
            val_pie, labels=None,
            autopct=lambda p: f"{p:.0f}%\n({int(round(p * tot_pie / 100))})",
            colors=col_pie, startangle=90, pctdistance=0.72,
            textprops={"fontsize": 9},
        )
        for at in autos_t:
            at.set_fontweight("bold")
        ax_tc.legend(wedges_t, [f"{l}  ({v})" for l, v in zip(lbl_pie, val_pie)],
                     loc="center left", bbox_to_anchor=(1, 0.5), fontsize=8)
        ax_tc.set_title("COLOCACIÓN DE TERCEROS", fontsize=13, fontweight="bold", pad=10)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_tc), f"A{cs_tc}", IMG_HALF_W, IMG_H)
        plt.close(fig_tc)

    if "Marca_Tercero" in df_f.columns:
        from collections import Counter
        df_mr = df_f[(df_f["Colocacion_Terceros"] == "Sí") &
                     (df_f["Marca_Tercero"].astype(str).str.strip().str.lower() != "nan") &
                     (df_f["Marca_Tercero"].astype(str).str.strip() != "")]["Marca_Tercero"].copy()
        todas_m = []
        for v in df_mr:
            for m in str(v).split(","):
                mc = m.strip().upper()
                if mc and mc != "NAN":
                    todas_m.append(mc)
        if todas_m:
            cteo   = Counter(todas_m)
            df_mrc = pd.DataFrame(cteo.items(), columns=["Marca", "Cantidad"]).sort_values("Cantidad")
            tot_m  = df_mrc["Cantidad"].sum()
            df_mrc["Pct"] = (df_mrc["Cantidad"] / tot_m * 100).round(1)
            fig_mr, ax_mr = plt.subplots(figsize=(9, max(3.5, len(df_mrc) * 0.55 + 1.5)))
            bars_mr = ax_mr.barh(df_mrc["Marca"], df_mrc["Cantidad"],
                                 color=[PALETTE[i % len(PALETTE)] for i in range(len(df_mrc))],
                                 edgecolor="white")
            max_mr = df_mrc["Cantidad"].max() if not df_mrc.empty else 1
            for bar, cnt, pct in zip(bars_mr, df_mrc["Cantidad"], df_mrc["Pct"]):
                ax_mr.text(bar.get_width() + max_mr * 0.03,
                           bar.get_y() + bar.get_height() / 2,
                           f"{cnt}\n({pct}%)", va="center", ha="left", fontsize=9, fontweight="bold")
            chart_style(ax_mr, "MARCAS DE COMPETENCIA EN PDV",
                        xlabel="Número de Visitas con Presencia")
            ax_mr.set_xlim(0, max_mr * 1.45)
            ax_mr.xaxis.grid(True, linestyle="--", alpha=0.45)
            plt.tight_layout()
            insert_image_bytes(fig_to_bytes(fig_mr), f"G{cs_tc}", IMG_HALF_W,
                               int(max(3.5, len(df_mrc) * 0.55 + 1.5) * 72))
            plt.close(fig_mr)

    row_cur += 1

    # CONTAMINACIÓN + VISIBILIDAD
    row_cur = section_header(row_cur, "CONTAMINACIÓN DE EXHIBIDORES  |  VISIBILIDAD POR EXHIBIDOR", "e05252")
    cs_cv   = row_cur
    row_cur = reserve_rows(row_cur, 24, 15)

    cont_info = [("CONT_LEGOS_GC", "LEGOS G&C"),
                 ("CONT_TOBOGAN_RITZ_OREO", "TOBOGÁN Ritz/Oreo"),
                 ("CONT_EXHIB_KIWI", "EXHIB KIWI")]
    c_lbls, c_si, c_no, p_si, p_no = [], [], [], [], []
    for col_c, lbl_c in cont_info:
        if col_c in df_f.columns:
            sc  = pd.to_numeric(df_f[col_c], errors="coerce").fillna(0)
            si  = int(sc.sum())
            no  = len(sc) - si
            c_lbls.append(lbl_c); c_si.append(si); c_no.append(no)
            p_si.append(round(si / len(sc) * 100, 1) if len(sc) else 0)
            p_no.append(round(no / len(sc) * 100, 1) if len(sc) else 0)

    if c_lbls:
        xc   = np.arange(len(c_lbls))
        wc2  = 0.35
        maxc = max(c_si + c_no) if c_si + c_no else 1
        fig_ct, ax_ct = plt.subplots(figsize=(9, 5))
        b1 = ax_ct.bar(xc - wc2/2, c_si, wc2, label="Contaminado", color="#e05252", edgecolor="white")
        b2 = ax_ct.bar(xc + wc2/2, c_no, wc2, label="Limpio",      color="#4CAF50", edgecolor="white")
        for bar, cnt, pct in zip(b1, c_si, p_si):
            if cnt > 0:
                ax_ct.text(bar.get_x() + bar.get_width()/2,
                           bar.get_height() + maxc * 0.025,
                           f"{cnt}\n({pct}%)", ha="center", va="bottom", fontsize=10, fontweight="bold")
        for bar, cnt, pct in zip(b2, c_no, p_no):
            if cnt > 0:
                ax_ct.text(bar.get_x() + bar.get_width()/2,
                           bar.get_height() + maxc * 0.025,
                           f"{cnt}\n({pct}%)", ha="center", va="bottom", fontsize=10, fontweight="bold")
        ax_ct.set_xticks(xc)
        ax_ct.set_xticklabels(c_lbls, fontsize=10)
        chart_style(ax_ct, "CONTAMINACIÓN DE EXHIBIDORES", ylabel="Cantidad de Visitas")
        ax_ct.set_ylim(0, maxc * 1.45)
        ax_ct.legend(fontsize=10)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_ct), f"A{cs_cv}", IMG_HALF_W, IMG_H)
        plt.close(fig_ct)

    vis_info = [("Visibilidad_Legos",   "LEGOS G&C"),
                ("Visibilidad_Tobogan", "TOBOGÁN Ritz/Oreo"),
                ("Visibilidad_Kiwi",    "EXHIB KIWI"),
                ("Visibilidad_Otros",   "OTROS")]
    v_alta, v_med, v_baj, v_lbls = [], [], [], []
    for col_v, lbl_v in vis_info:
        if col_v in df_f.columns:
            sv = pd.to_numeric(df_f[col_v], errors="coerce").fillna(0)
            v_alta.append(int((sv == 1).sum()))
            v_med.append(int((sv == 2).sum()))
            v_baj.append(int((sv == 3).sum()))
            v_lbls.append(lbl_v)

    if v_lbls:
        xv   = np.arange(len(v_lbls))
        wv   = 0.25
        maxv = max(v_alta + v_med + v_baj) if v_alta + v_med + v_baj else 1
        fig_vi, ax_vi = plt.subplots(figsize=(10, 5))
        ba = ax_vi.bar(xv - wv, v_alta, wv, label="Alta",  color="#4CAF50", edgecolor="white")
        bm = ax_vi.bar(xv,      v_med,  wv, label="Media", color="#FFC107", edgecolor="white")
        bb = ax_vi.bar(xv + wv, v_baj,  wv, label="Baja",  color="#e05252", edgecolor="white")
        for bars_v, data_v in [(ba, v_alta), (bm, v_med), (bb, v_baj)]:
            for bar, cnt in zip(bars_v, data_v):
                if cnt > 0:
                    pct_v = round(cnt / total_visitas * 100, 1)
                    ax_vi.text(bar.get_x() + bar.get_width()/2,
                               bar.get_height() + maxv * 0.025,
                               f"{cnt}\n({pct_v}%)",
                               ha="center", va="bottom", fontsize=8, fontweight="bold")
        ax_vi.set_xticks(xv)
        ax_vi.set_xticklabels(v_lbls, fontsize=10)
        chart_style(ax_vi, "VISIBILIDAD POR EXHIBIDOR", ylabel="Cantidad de Visitas")
        ax_vi.set_ylim(0, maxv * 1.5)
        ax_vi.legend(title="Visibilidad", fontsize=10)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_vi), f"G{cs_cv}", IMG_HALF_W, IMG_H)
        plt.close(fig_vi)

    row_cur += 1

    # PRESENCIA DE PRODUCTOS
    productos_grupos = {
        "BISCUITS": {
            "OREO_34GR": "OREO 34GR", "OREO_54GR": "OREO 54GR", "OREO_ROLLO": "OREO ROLLO",
            "RITZ_ROLLO": "RITZ ROLLO", "RITZ_PACK": "RITZ PACK",
            "FIELD_CC": "FIELD CC", "FIELD_DP": "FIELD DP", "FIELD_VAIN": "FIELD VAIN",
            "CLUB_SOCIAL_TRA": "CLUB SOCIAL TRA",
        },
        "PRODUCTOS FOCO": {
            "OREO_FRESA_PACK": "OREO FRESA (Pack)", "OREO_FRESA_ROLLO": "OREO FRESA (Rollo)",
            "OREO_CHOCO_LIMON_PACK": "OREO CHOCO LIMÓN (Pack)",
            "OREO_CHOCO_LIMON_ROLLO": "OREO CHOCO LIMÓN (Rollo)",
            "CLUB_SOCIAL_SAB": "CLUB SOCIAL (Sabores)",
            "ROLLO_GOLDEN": "OREO GOLDEN (Rollo)", "ROLLO_CHOCOLATE": "OREO CHOCOLATE (Rollo)",
        },
        "G&C": {
            "TRIDENT_5s": "TRIDENT 5s", "TRIDENT_EVUP": "TRIDENT EVUP",
            "HALLS_12s": "HALLS 12s", "HALLS_100s": "HALLS 100s",
            "CHICLETS_2S": "CHICLETS 2S", "BUBBALOO": "BUBBALOO",
        },
    }

    row_cur = section_header(row_cur, "PRESENCIA DE PRODUCTOS (% DE VISITAS)", "1a1a1a")

    for gp_name, prods in productos_grupos.items():
        cs_pr   = row_cur
        row_cur = reserve_rows(row_cur, 22, 15)
        pres_d  = []
        for kp, lp in prods.items():
            if kp in df_f.columns:
                sp  = pd.to_numeric(df_f[kp], errors="coerce")
                cnt = int(sp.sum())
                pct = round(sp.mean() * 100, 1)
                pres_d.append({"Producto": lp, "Pct": pct, "Cnt": cnt})
        if not pres_d:
            row_cur += 1
            continue
        df_pr = pd.DataFrame(pres_d).sort_values("Pct", ascending=False)
        fig_pr, ax_pr = plt.subplots(figsize=(13, 5))
        bars_pr = ax_pr.bar(df_pr["Producto"], df_pr["Pct"],
                            color=[PALETTE[i % len(PALETTE)] for i in range(len(df_pr))],
                            edgecolor="white", linewidth=1, zorder=3)
        for bar, pct, cnt in zip(bars_pr, df_pr["Pct"], df_pr["Cnt"]):
            ax_pr.text(bar.get_x() + bar.get_width()/2,
                       bar.get_height() + 1.8,
                       f"{pct}%\n({cnt}/{total_visitas})",
                       ha="center", va="bottom", fontsize=9, fontweight="bold")
        chart_style(ax_pr, f"PRESENCIA DE PRODUCTOS — {gp_name}", ylabel="Presencia %")
        ax_pr.set_ylim(0, 118)
        plt.xticks(rotation=22, ha="right", fontsize=9)
        plt.tight_layout()
        insert_image_bytes(fig_to_bytes(fig_pr), f"A{cs_pr}", IMG_FULL_W, IMG_H_WIDE)
        plt.close(fig_pr)
        row_cur += 1

    # TABLA ÚLTIMAS VISITAS
    row_cur = section_header(row_cur, "ÚLTIMAS VISITAS — DETALLE", "1a1a1a")
    cols_tbl = ["Fecha", "Codigo_PDC", "Nombre_Cliente", "Giro_Negocio", "Vendedor",
                "Zona", "Efectividad_Soles", "Tiempo_PDC",
                "Colocacion_Terceros", "Marca_Tercero"]
    cols_ex  = [c for c in cols_tbl if c in df_f.columns]
    df_tbl   = df_f[cols_ex].sort_values("Fecha", ascending=False).head(100)

    hdr_cols = ["4472C4", "4472C4", "4472C4", "6a9e4f", "6a9e4f",
                "FF7F0E", "7b5ea7", "e05252", "FFC107", "FFC107"]
    for ci, col_n in enumerate(cols_ex, 1):
        c = ws.cell(row=row_cur, column=ci, value=col_n.replace("_", " ").upper())
        hc = hdr_cols[ci - 1] if ci - 1 < len(hdr_cols) else "4472C4"
        c.font      = Font(bold=True, size=10, color="FFFFFF", name="Arial")
        c.fill      = PatternFill("solid", fgColor=hc)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    ws.row_dimensions[row_cur].height = 22
    row_cur += 1

    alt_fill2 = PatternFill("solid", fgColor="F5F5F5")
    for ri, (_, rd) in enumerate(df_tbl.iterrows()):
        for ci, col_n in enumerate(cols_ex, 1):
            val = rd[col_n]
            if pd.isna(val): val = ""
            c = ws.cell(row=row_cur, column=ci, value=val)
            c.font      = Font(size=10, name="Arial")
            c.border    = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 1:
                c.fill = alt_fill2
        ws.row_dimensions[row_cur].height = 16
        row_cur += 1

    for col_l, w in {"A": 32, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18,
                     "G": 32, "H": 18, "I": 18, "J": 18, "K": 18, "L": 18}.items():
        ws.column_dimensions[col_l].width = w

    ws_raw  = wb.create_sheet("Datos Completos")
    exp_df  = df_f.copy()
    exp_df  = exp_df.merge(ticket_calc[["Vendedor", "Fecha_str", "Ticket_Calculado"]],
                           on=["Vendedor", "Fecha_str"], how="left")
    cols_rw = [c for c in exp_df.columns
               if c not in ["Imagen_Path", "Fecha_str", "Concreto"]]
    for ci, col_n in enumerate(cols_rw, 1):
        c = ws_raw.cell(row=1, column=ci, value=col_n)
        c.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        c.fill      = PatternFill("solid", fgColor="1a1a1a")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_raw.column_dimensions[get_column_letter(ci)].width = 18
    for ri, (_, rd) in enumerate(exp_df[cols_rw].iterrows(), 2):
        for ci, val in enumerate(rd, 1):
            if pd.isna(val): val = ""
            ws_raw.cell(row=ri, column=ci, value=val).font = Font(size=10, name="Arial")

    wb.save(dash_buf)
    dash_buf.seek(0)
    return dash_buf


# ═══════════════════════════════════════════
# PÁGINA: FORMULARIO
# ═══════════════════════════════════════════
if st.session_state.pagina == "formulario":

    df_check = cargar_datos()
    if not df_check.empty:
        col_nav1, col_nav2 = st.columns([6, 1])
        with col_nav2:
            if st.button("📊 Ver Dashboard"):
                st.session_state.pagina = "dashboard"
                st.rerun()

    st.markdown("# Registro de Visita")
    st.markdown("---")

    st.markdown("### 📍 Ubicación del PDC")

    if st.session_state.gps_lat and st.session_state.gps_lon:
        lugar_guardado = st.session_state.get("gps_lugar", "")
        if lugar_guardado:
            st.success(f"✅ Ubicación guardada: {lugar_guardado} | Coordenadas: {st.session_state.gps_lat}, {st.session_state.gps_lon}")
        else:
            st.success(f"✅ Ubicación guardada: {st.session_state.gps_lat}, {st.session_state.gps_lon}")
        if st.button("Cambiar ubicación"):
            st.session_state.gps_lat = ""
            st.session_state.gps_lon = ""
            st.session_state.geo_resultados = []
            if "map_click" in st.session_state:
                del st.session_state["map_click"]
            st.rerun()
    else:
        if "ajuste_manual" not in st.session_state:
            st.session_state["ajuste_manual"] = False

        col_tog1, col_tog2 = st.columns([1, 3])
        with col_tog1:
            if st.session_state["ajuste_manual"]:
                if st.button("🔒 Desactivar ajuste", key="btn_toggle_ajuste", use_container_width=True):
                    st.session_state["ajuste_manual"] = False
                    st.rerun()
            else:
                if st.button("✏️ Ajuste manual", key="btn_toggle_ajuste", use_container_width=True):
                    st.session_state["ajuste_manual"] = True
                    st.rerun()
        with col_tog2:
            if st.session_state["ajuste_manual"]:
                st.success("✅ Modo ajuste activo — mueve el mapa hasta centrar el pin rojo en el PDC y presiona 'Obtener ubicación'")
            else:
                st.info("ℹ️ Activa el ajuste manual para mover el marcador al PDC")

        center = st.session_state.get("map_center", [-8.111801, -79.028678])
        zoom   = st.session_state.get("map_zoom", 15)
        m = folium.Map(location=center, zoom_start=zoom, tiles="OpenStreetMap")

        if not st.session_state["ajuste_manual"]:
            folium.Marker(
                location=center, draggable=False,
                tooltip="📍 Activa el ajuste manual para ubicar el PDC",
                icon=folium.Icon(color="gray", icon="map-marker", prefix="fa"),
            ).add_to(m)

        if st.session_state["ajuste_manual"]:
            m.get_root().html.add_child(folium.Element("""
            <style>
            .crosshair-shadow { position:absolute;top:50%;left:50%;transform:translate(-50%,2px);width:14px;height:6px;background:rgba(0,0,0,0.25);border-radius:50%;z-index:9999;pointer-events:none; }
            .crosshair-pin { position:absolute;top:50%;left:50%;transform:translate(-50%,-100%);z-index:9999;pointer-events:none;animation:bounce 1.2s infinite ease-in-out; }
            .crosshair-lines { position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);width:60px;height:60px;z-index:9998;pointer-events:none; }
            @keyframes bounce { 0%,100%{transform:translate(-50%,-100%)}50%{transform:translate(-50%,-115%)} }
            </style>
            <div class="crosshair-shadow"></div>
            <div class="crosshair-pin"><svg width="36" height="48" viewBox="0 0 36 48" xmlns="http://www.w3.org/2000/svg"><filter id="shadow"><feDropShadow dx="0" dy="2" stdDeviation="2" flood-color="rgba(0,0,0,0.3)"/></filter><path d="M18 0C8.059 0 0 8.059 0 18c0 12 18 30 18 30S36 30 36 18C36 8.059 27.941 0 18 0z" fill="#e05252" filter="url(#shadow)"/><circle cx="18" cy="18" r="8" fill="white"/><circle cx="18" cy="18" r="4.5" fill="#e05252"/></svg></div>
            <svg class="crosshair-lines" viewBox="0 0 60 60" xmlns="http://www.w3.org/2000/svg"><line x1="0" y1="30" x2="22" y2="30" stroke="#e05252" stroke-width="1.5" stroke-dasharray="4 3" opacity="0.6"/><line x1="38" y1="30" x2="60" y2="30" stroke="#e05252" stroke-width="1.5" stroke-dasharray="4 3" opacity="0.6"/><line x1="30" y1="0" x2="30" y2="22" stroke="#e05252" stroke-width="1.5" stroke-dasharray="4 3" opacity="0.6"/><line x1="30" y1="38" x2="30" y2="60" stroke="#e05252" stroke-width="1.5" stroke-dasharray="4 3" opacity="0.6"/></svg>
            """))

        map_data = st_folium(m, height=420, use_container_width=True, key="folium_map", returned_objects=["center", "zoom"])

        if map_data and map_data.get("center"):
            c = map_data["center"]
            st.session_state["map_center_actual"] = [round(c["lat"], 6), round(c["lng"], 6)]

        if st.session_state["ajuste_manual"]:
            col_ob1, col_ob2 = st.columns([2, 3])
            with col_ob1:
                st.caption("Mueve el mapa hasta centrar el punto rojo en el PDC")
            with col_ob2:
                if st.button("📌 Obtener ubicación del centro", key="btn_obtener_ubi", use_container_width=True):
                    pos = st.session_state.get("map_center_actual") or center
                    lat_nuevo = round(pos[0], 6)
                    lon_nuevo = round(pos[1], 6)
                    st.session_state["map_click"] = (lat_nuevo, lon_nuevo)
                    st.session_state["map_center"] = [lat_nuevo, lon_nuevo]
                    with st.spinner("Obteniendo nombre del lugar..."):
                        try:
                            import requests as _req
                            r = _req.get("https://nominatim.openstreetmap.org/reverse",
                                params={"lat": lat_nuevo, "lon": lon_nuevo, "format": "json", "accept-language": "es", "zoom": 18},
                                headers={"User-Agent": "MDZ-SupervisionApp/1.0"}, timeout=6)
                            if r.status_code == 200:
                                addr = r.json().get("address", {})
                                nombre = (addr.get("road") or addr.get("pedestrian") or addr.get("neighbourhood") or addr.get("suburb") or addr.get("town") or addr.get("city") or "Ubicación seleccionada")
                                ciudad = addr.get("city") or addr.get("town") or addr.get("village") or addr.get("county") or ""
                                distrito = addr.get("suburb") or addr.get("neighbourhood") or ""
                                partes = [p for p in [distrito, ciudad] if p and p != nombre]
                                st.session_state["map_lugar"] = ", ".join([nombre] + partes[:2])
                            else:
                                st.session_state["map_lugar"] = ""
                        except Exception:
                            st.session_state["map_lugar"] = ""
                    st.rerun()

        if st.session_state.get("map_click"):
            lat_click, lon_click = st.session_state["map_click"]
            lugar = st.session_state.get("map_lugar", "")
            st.markdown("---")
            col_info1, col_info2 = st.columns([3, 1])
            with col_info1:
                if lugar:
                    st.markdown(f"""<div style="background:#f0f4ff;border:1px solid #c5d0e8;border-radius:8px;padding:10px 14px;">
                        <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px;">Ubicación seleccionada</div>
                        <div style="font-size:15px;font-weight:600;color:#1a1a1a;">📍 {lugar}</div>
                        <div style="font-size:12px;color:#888;margin-top:2px;">{lat_click}, {lon_click}</div>
                    </div>""", unsafe_allow_html=True)
                else:
                    st.markdown(f"""<div style="background:#f0f4ff;border:1px solid #c5d0e8;border-radius:8px;padding:10px 14px;">
                        <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px;">Ubicación seleccionada</div>
                        <div style="font-size:15px;font-weight:600;color:#1a1a1a;">📍 {lat_click}, {lon_click}</div>
                    </div>""", unsafe_allow_html=True)
            with col_info2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("✅ Confirmar punto", key="geo_map_ok", use_container_width=True):
                    st.session_state.gps_lat = str(lat_click)
                    st.session_state.gps_lon = str(lon_click)
                    lugar_final = st.session_state.get("map_lugar", "")
                    if not lugar_final:
                        with st.spinner("Obteniendo nombre..."):
                            try:
                                import requests as _req
                                r = _req.get("https://nominatim.openstreetmap.org/reverse",
                                    params={"lat": lat_click, "lon": lon_click, "format": "json", "accept-language": "es", "zoom": 18},
                                    headers={"User-Agent": "MDZ-SupervisionApp/1.0"}, timeout=6)
                                if r.status_code == 200:
                                    addr = r.json().get("address", {})
                                    nombre = (addr.get("road") or addr.get("pedestrian") or addr.get("neighbourhood") or addr.get("suburb") or addr.get("town") or addr.get("city") or "Ubicación seleccionada")
                                    ciudad = addr.get("city") or addr.get("town") or addr.get("village") or addr.get("county") or ""
                                    distrito = addr.get("suburb") or addr.get("neighbourhood") or ""
                                    partes = [p for p in [nombre, distrito, ciudad] if p and p != nombre]
                                    lugar_final = ", ".join([nombre] + partes[:2])
                            except Exception:
                                lugar_final = ""
                    st.session_state["gps_lugar"] = lugar_final
                    st.session_state["ajuste_manual"] = False
                    st.session_state.pop("map_click", None)
                    st.session_state.pop("map_lugar", None)
                    st.rerun()

    st.markdown("---")

    with st.form("form_visita", clear_on_submit=False):
        st.markdown("### 🏪 Datos del Cliente")
        col1, col2, col3 = st.columns(3)
        with col1: fecha = st.date_input("Fecha de Visita", value=date.today())
        with col2: codigo_pdc = st.text_input("Código PDC (8 dígitos)", max_chars=8, placeholder="Ej: 00000001")
        with col3: nombre_cliente = st.text_input("Nombre del Cliente", placeholder="Ej: Bodega Central")

        col4, col5 = st.columns(2)
        with col4:
            giro_negocio = st.selectbox("Giro de Negocio", [
                "Selecciona...", "1 - Bodega", "2 - Minimarket / Tiendas", "3 - Kiosko",
                "4 - Especializados (Panificadora, Horeca, Internet...)",
                "5 - Otros (Puesto de mercado, Centros Educativos...)"
            ])
        with col5:
            zona = st.text_input("Zona", placeholder="Ej: TRUJILLO CENTRO, VICTOR LARCO...")

        latitud  = st.session_state.gps_lat
        longitud = st.session_state.gps_lon

        st.markdown("---")
        st.markdown("### 🧑‍💼 Ruta")
        col_v1, col_v2, col_v3 = st.columns(3)
        with col_v1:
            vendedor = st.text_input("Nombre del Vendedor",
                placeholder="Escribe el nombre completo del vendedor", key="txt_vendedor")
        with col_v2:
            codigo_vendedor = st.text_input("Código de Vendedor", max_chars=8, placeholder="Ej: VEN00001")
        with col_v3:
            mesa = st.text_input("Mesa", placeholder="Ej: DJ1, DJ3...", key="txt_mesa")
        ruta_logica = st.text_input("Ruta Lógica", placeholder="Ej: Ruta 01 - Norte")

        st.markdown("---")
        st.markdown("### 🍪 Presencia Biscuits")
        biscuits = {
            "OREO_34GR": "OREO 34GR", "OREO_54GR": "OREO 54GR", "OREO_ROLLO": "OREO ROLLO",
            "RITZ_ROLLO": "RITZ ROLLO", "RITZ_PACK": "RITZ PACK",
            "FIELD_CC": "FIELD (CC)", "FIELD_DP": "FIELD (DP)", "FIELD_VAIN": "FIELD (VAIN)",
            "CLUB_SOCIAL_TRA": "CLUB SOCIAL (TRA)",
        }
        cols_b = st.columns(5)
        biscuits_vals = {}
        for i, (key, label) in enumerate(biscuits.items()):
            with cols_b[i % 5]:
                biscuits_vals[key] = st.checkbox(label, key=f"b_{key}")

        st.markdown("---")
        st.markdown("### ⭐ Productos Foco")
        productos_foco = {
            "OREO_FRESA_PACK":        "OREO FRESA (Pack)",
            "OREO_FRESA_ROLLO":       "OREO FRESA (Rollo)",
            "OREO_CHOCO_LIMON_PACK":  "OREO CHOCO LIMÓN (Pack)",
            "OREO_CHOCO_LIMON_ROLLO": "OREO CHOCO LIMÓN (Rollo)",
            "CLUB_SOCIAL_SAB":        "CLUB SOCIAL (Sabores)",
            "ROLLO_GOLDEN":           "OREO GOLDEN (Rollo)",
            "ROLLO_CHOCOLATE":        "OREO CHOCOLATE (Rollo)",
        }
        cols_pf = st.columns(3)
        pf_vals = {}
        for i, (key, label) in enumerate(productos_foco.items()):
            with cols_pf[i % 3]:
                pf_vals[key] = st.checkbox(label, key=f"pf_{key}")

        st.markdown("---")
        st.markdown("### 🍬 Presencia G&C")
        gyc = {
            "TRIDENT_5s": "TRIDENT 5s", "TRIDENT_EVUP": "TRIDENT EVUP",
            "HALLS_12s": "HALLS 12s", "HALLS_100s": "HALLS 100s",
            "CHICLETS_2S": "CHICLETS 2S", "BUBBALOO": "BUBBALOO",
        }
        cols2 = st.columns(6)
        gyc_vals = {}
        for i, (key, label) in enumerate(gyc.items()):
            with cols2[i % 6]:
                gyc_vals[key] = st.checkbox(label, key=f"g_{key}")

        st.markdown("---")
        st.markdown("### 🗂️ Tipos de Exhibidores")
        tipos = {
            "LEGOS_GC": "LEGOS G&C", "TOBOGAN_RITZ_OREO": "TOBOGÁN (Ritz/Oreo)",
            "EXHIB_KIWI": "EXHIB KIWI", "RITRAZ": "RITRAZ",
            "MEGA_KIWI": "MEGA KIWI", "EXHIBIDOR_OTROS": "OTROS",
        }
        cols3 = st.columns(3)
        tipos_vals = {}
        for i, (key, label) in enumerate(tipos.items()):
            with cols3[i % 3]:
                tipos_vals[key] = st.checkbox(label, key=f"t_{key}")
        exhibidor_otros_desc = st.text_input("Especificar otro exhibidor (si marcó OTROS)", placeholder="Ej: Stand especial, Canastilla...", key="exhib_otros_desc")

        st.markdown("---")
        st.markdown("### ⚠️ Contaminación de Exhibidores")
        cols_cont = st.columns(3)
        with cols_cont[0]: cont_legos   = st.radio("LEGOS G&C", ["No", "Sí"], horizontal=True, key="cr_legos")
        with cols_cont[1]: cont_tobogan = st.radio("TOBOGÁN (Ritz/Oreo)", ["No", "Sí"], horizontal=True, key="cr_tobogan")
        with cols_cont[2]: cont_kiwi    = st.radio("EXHIB KIWI", ["No", "Sí"], horizontal=True, key="cr_kiwi")
        causa_contaminacion = ""
        if cont_legos == "Sí" or cont_tobogan == "Sí" or cont_kiwi == "Sí":
            causa_contaminacion = st.text_input("Causa de contaminación", placeholder="Describe la causa...", key="causa_cont")
        cont_vals = {
            "CONT_LEGOS_GC":          1 if cont_legos   == "Sí" else 0,
            "CONT_TOBOGAN_RITZ_OREO": 1 if cont_tobogan == "Sí" else 0,
            "CONT_EXHIB_KIWI":        1 if cont_kiwi    == "Sí" else 0,
        }

        st.markdown("---")
        st.markdown("### 👁️ Visibilidad por Exhibidor")
        st.markdown('<div class="leyenda-box">1 = Alta &nbsp;|&nbsp; 2 = Media &nbsp;|&nbsp; 3 = Baja</div>', unsafe_allow_html=True)
        VIS_OPTIONS = [1, 2, 3]
        VIS_LABELS  = {1: "1 - Alta", 2: "2 - Media", 3: "3 - Baja"}
        v1, v2, v3, v4 = st.columns(4)
        with v1: vis_legos   = st.radio("LEGOS G&C",           VIS_OPTIONS, format_func=lambda x: VIS_LABELS[x], horizontal=True, key="vl")
        with v2: vis_tobogan = st.radio("TOBOGÁN (Ritz/Oreo)", VIS_OPTIONS, format_func=lambda x: VIS_LABELS[x], horizontal=True, key="vt")
        with v3: vis_kiwi    = st.radio("EXHIB KIWI",          VIS_OPTIONS, format_func=lambda x: VIS_LABELS[x], horizontal=True, key="vk")
        with v4: vis_otros   = st.radio("OTROS",               VIS_OPTIONS, format_func=lambda x: VIS_LABELS[x], horizontal=True, key="vo")
        vis_otros_desc = st.text_input("Descripción de OTROS (visibilidad)", placeholder="Ej: Exhibidor especial...", key="vis_otros_desc")

        st.markdown("---")
        st.markdown("### 🏷️ Colocación de Terceros")
        col_terc1, col_terc2 = st.columns([1, 3])
        with col_terc1:
            colocacion_terceros = st.radio("¿Hay colocación de terceros?", ["No", "Sí"], horizontal=True)
        with col_terc2:
            t1, t2, t3, t4 = st.columns(4)
            with t1: marca_tercero_1 = st.text_input("Marca 1", placeholder="Ej: Alicorp", key="mt1")
            with t2: marca_tercero_2 = st.text_input("Marca 2", placeholder="Ej: Molitalia", key="mt2")
            with t3: marca_tercero_3 = st.text_input("Marca 3", placeholder="Ej: Costa", key="mt3")
            with t4: marca_tercero_4 = st.text_input("Marca 4", placeholder="Ej: Nestlé", key="mt4")
        marca_tercero = ", ".join([m for m in [marca_tercero_1, marca_tercero_2, marca_tercero_3, marca_tercero_4] if m.strip()])

        st.markdown("---")
        st.markdown("### 📊 Indicadores de la visita")
        k1, k2 = st.columns(2)
        with k1: efectividad_soles = st.number_input("Efectividad (S/)", min_value=0.0, step=1.0, value=0.0, format="%.2f")
        with k2: tiempo_pdc = st.number_input("Tiempo en PDC (minutos)", min_value=0, step=1, value=0)

        st.markdown("---")
        st.markdown("### 📝 Detalles IG")
        detalles_ig = st.text_area("Detalles IG", placeholder="Ingresa los detalles IG...", height=100, label_visibility="collapsed")

        st.markdown("---")
        st.markdown("### 📷 Evidencia fotográfica")
        st.caption("Puedes subir hasta 5 fotos por visita (JPG, PNG)")
        imagen_1 = st.file_uploader("Foto 1", type=["jpg", "jpeg", "png"], key="img1")
        imagen_2 = st.file_uploader("Foto 2", type=["jpg", "jpeg", "png"], key="img2")
        imagen_3 = st.file_uploader("Foto 3", type=["jpg", "jpeg", "png"], key="img3")
        imagen_4 = st.file_uploader("Foto 4", type=["jpg", "jpeg", "png"], key="img4")
        imagen_5 = st.file_uploader("Foto 5", type=["jpg", "jpeg", "png"], key="img5")

        st.markdown("---")
        submitted = st.form_submit_button("💾 Guardar registro", use_container_width=True)

        if submitted:
            if not codigo_pdc or not nombre_cliente or giro_negocio == "Selecciona...":
                st.error("Por favor completa el Código PDC, Nombre del Cliente y Giro de Negocio.")
            else:
                img_paths = []
                nombre_limpio = nombre_cliente.strip().replace(" ", "_").replace("/", "-")
                for idx, img_file in enumerate([imagen_1, imagen_2, imagen_3, imagen_4, imagen_5], start=1):
                    if img_file is not None:
                        ext = img_file.name.rsplit(".", 1)[-1].lower()
                        img_filename = f"{nombre_limpio}_{str(fecha)}_foto{idx}.{ext}"
                        img_path_full = os.path.join(IMG_FOLDER, img_filename)
                        with open(img_path_full, "wb") as fout:
                            fout.write(img_file.getbuffer())
                        img_paths.append(img_path_full)

                img_path_guardado = "|".join(img_paths) if img_paths else ""
                registro = {
                    "Fecha": str(fecha), "Codigo_PDC": codigo_pdc,
                    "Nombre_Cliente": nombre_cliente, "Giro_Negocio": giro_negocio,
                    "Vendedor": vendedor, "Codigo_Vendedor": codigo_vendedor,
                    "Mesa": mesa, "Zona": zona, "Latitud": latitud, "Longitud": longitud,
                    **{k: int(v) for k, v in biscuits_vals.items()},
                    **{k: int(v) for k, v in pf_vals.items()},
                    **{k: int(v) for k, v in gyc_vals.items()},
                    **{k: int(v) for k, v in tipos_vals.items()},
                    "EXHIBIDOR_OTROS_DESC": exhibidor_otros_desc,
                    **{k: int(v) for k, v in cont_vals.items()},
                    "Causa_Contaminacion": causa_contaminacion,
                    "Visibilidad_Legos":   vis_legos,
                    "Visibilidad_Tobogan": vis_tobogan,
                    "Visibilidad_Kiwi":    vis_kiwi,
                    "Visibilidad_Otros":   vis_otros,
                    "Visibilidad_Otros_Desc": vis_otros_desc,
                    "Colocacion_Terceros": colocacion_terceros,
                    "Marca_Tercero":       marca_tercero,
                    "Efectividad_Soles":   efectividad_soles,
                    "Tiempo_PDC":          tiempo_pdc,
                    "Imagen_Path":         img_path_guardado,
                }
                guardar_registro(registro)
                st.session_state.df_editado = None  # reset editable df on new record
                st.session_state.pagina = "dashboard"
                st.rerun()


# ═══════════════════════════════════════════
# PÁGINA: DASHBOARD
# ═══════════════════════════════════════════
elif st.session_state.pagina == "dashboard":

    df = cargar_datos()

    col_title, col_btn1, col_btn2, col_btn3 = st.columns([4, 1, 1, 1])
    with col_title:
        st.markdown("# Dashboard — Supervisión Canal Tradicional")
    with col_btn1:
        if st.button("← Ingresar datos"):
            st.session_state.pagina = "formulario"; st.rerun()
    with col_btn2:
        if st.button("＋ Nueva visita"):
            st.session_state.pagina = "formulario"; st.rerun()
    with col_btn3:
        if not st.session_state.confirmar_eliminar:
            st.markdown('<div class="btn-danger">', unsafe_allow_html=True)
            if st.button("🗑️ Eliminar historial"):
                st.session_state.confirmar_eliminar = True; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.warning("¿Seguro? Esta acción no se puede deshacer.")
            c1, c2 = st.columns(2)
            with c1:
                st.markdown('<div class="btn-danger">', unsafe_allow_html=True)
                if st.button("Confirmar eliminación"):
                    eliminar_historial()
                    st.session_state.confirmar_eliminar = False
                    st.session_state.df_editado = None
                    st.session_state.pagina = "formulario"; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with c2:
                if st.button("Cancelar"):
                    st.session_state.confirmar_eliminar = False; st.rerun()

    if df.empty:
        st.warning("No hay registros aún.")
        st.stop()

    df["Fecha"] = pd.to_datetime(df["Fecha"])
    for col in ["Efectividad_Soles", "Tiempo_PDC"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    st.markdown("---")
    st.markdown("#### 📅 Filtros y datos históricos")

    tab_actual, tab_historial = st.tabs(["📊 Datos actuales", "🗂️ Historial guardado"])

    with tab_actual:
        fecha_min = df["Fecha"].min().date()
        fecha_max = df["Fecha"].max().date()
        col_f1, col_f2, col_f3 = st.columns([2, 2, 3])
        with col_f1: fecha_desde = st.date_input("Desde", value=fecha_min, min_value=fecha_min, max_value=fecha_max, key="fd")
        with col_f2: fecha_hasta = st.date_input("Hasta", value=fecha_max, min_value=fecha_min, max_value=fecha_max, key="fh")
        with col_f3:
            vendedores_disponibles = ["Todos"] + sorted(df["Vendedor"].dropna().unique().tolist())
            filtro_vendedor = st.selectbox("Filtrar por Vendedor", vendedores_disponibles)

        st.markdown("---")
        st.markdown("##### 💾 Guardar este período como histórico")
        sc1, sc2 = st.columns([3, 1])
        with sc1:
            nombre_snap = st.text_input("Nombre del período",
                placeholder=f"Ej: Semana del {fecha_desde} al {fecha_hasta}",
                key="snap_nombre", label_visibility="collapsed")
        with sc2:
            if st.button("💾 Guardar período", use_container_width=True, key="btn_guardar_snap"):
                mask_snap = (df["Fecha"].dt.date >= fecha_desde) & (df["Fecha"].dt.date <= fecha_hasta)
                df_snap = df[mask_snap].copy()
                if filtro_vendedor != "Todos":
                    df_snap = df_snap[df_snap["Vendedor"] == filtro_vendedor]
                if df_snap.empty:
                    st.warning("No hay datos en este rango.")
                else:
                    snap_key = nombre_snap.strip() if nombre_snap.strip() else f"{fecha_desde} → {fecha_hasta}"
                    if filtro_vendedor != "Todos": snap_key += f" ({filtro_vendedor})"
                    st.session_state.snapshots[snap_key] = df_snap.to_csv(index=False)
                    guardar_snapshots_disco()
                    st.success(f"✅ Guardado: *{snap_key}* ({len(df_snap)} registros)")

    with tab_historial:
        if not st.session_state.snapshots:
            st.info("Aún no hay períodos guardados.")
        else:
            snap_names = list(st.session_state.snapshots.keys())
            snap_sel = st.selectbox("Selecciona un período guardado:", snap_names, key="snap_sel")
            hcol1, hcol2 = st.columns([1, 1])
            with hcol1:
                if st.button("📂 Ver este período", use_container_width=True, key="btn_ver_snap"):
                    st.session_state["snap_activo"] = snap_sel
            with hcol2:
                st.markdown('<div class="btn-danger">', unsafe_allow_html=True)
                if st.button("🗑️ Eliminar este período", use_container_width=True, key="btn_del_snap"):
                    del st.session_state.snapshots[snap_sel]
                    if st.session_state.get("snap_activo") == snap_sel:
                        del st.session_state["snap_activo"]
                    guardar_snapshots_disco(); st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

            if st.session_state.get("snap_activo") and st.session_state["snap_activo"] in st.session_state.snapshots:
                snap_activo = st.session_state["snap_activo"]
                df_snap_view = pd.read_csv(io.StringIO(st.session_state.snapshots[snap_activo]))
                df_snap_view["Fecha"] = pd.to_datetime(df_snap_view["Fecha"])
                st.success(f"📂 Mostrando: *{snap_activo}* — {len(df_snap_view)} registros")
                for col2 in ["Efectividad_Soles", "Tiempo_PDC"]:
                    if col2 in df_snap_view.columns:
                        df_snap_view[col2] = pd.to_numeric(df_snap_view[col2], errors="coerce").fillna(0)
                sv1, sv2, sv3, sv4 = st.columns(4)
                snap_ventas   = df_snap_view["Efectividad_Soles"].sum() if "Efectividad_Soles" in df_snap_view.columns else 0
                snap_visitas  = len(df_snap_view)
                snap_tiempo   = df_snap_view["Tiempo_PDC"].mean() if "Tiempo_PDC" in df_snap_view.columns else 0
                snap_terceros = (df_snap_view["Colocacion_Terceros"] == "Sí").mean() * 100 if "Colocacion_Terceros" in df_snap_view.columns else 0
                for col_s, lbl, val, sub in [
                    (sv1, "Visitas",      str(snap_visitas),        "registros"),
                    (sv2, "Efectividad",  f"S/ {snap_ventas:,.2f}", "total ventas"),
                    (sv3, "Tiempo Prom.", f"{snap_tiempo:.0f} min", "por visita"),
                    (sv4, "Con Terceros", f"{snap_terceros:.0f}%",  "de visitas"),
                ]:
                    with col_s:
                        st.markdown(f"""<div class="kpi-box"><div class="kpi-label">{lbl}</div>
                            <div class="kpi-value">{val}</div><div class="kpi-sub">{sub}</div></div>""",
                            unsafe_allow_html=True)
                st.markdown("")
                cols_snap = [c for c in ["Fecha","Codigo_PDC","Nombre_Cliente","Giro_Negocio",
                    "Vendedor","Zona","Efectividad_Soles","Tiempo_PDC","Colocacion_Terceros","Marca_Tercero"]
                    if c in df_snap_view.columns]
                st.dataframe(df_snap_view[cols_snap].sort_values("Fecha", ascending=False), use_container_width=True, hide_index=True)
                snap_buf = io.BytesIO()
                with pd.ExcelWriter(snap_buf, engine="openpyxl") as writer:
                    df_snap_view.drop(columns=["Imagen_Path","Concreto","Fecha_str"], errors="ignore").to_excel(writer, index=False, sheet_name="Historico")
                snap_buf.seek(0)
                st.download_button(label=f"⬇️ Descargar Excel — {snap_activo}", data=snap_buf,
                    file_name=f"historico_{snap_activo.replace(' ','_').replace('→','a')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    st.markdown("---")

    mask = (df["Fecha"].dt.date >= fecha_desde) & (df["Fecha"].dt.date <= fecha_hasta)
    df_f_original = df[mask].copy()
    if filtro_vendedor != "Todos":
        df_f_original = df_f_original[df_f_original["Vendedor"] == filtro_vendedor]
    if df_f_original.empty:
        st.warning("No hay registros en el rango seleccionado.")
        st.stop()

    # ── TABLA EDITABLE — gestión del estado ────────────────────────────────
    # Si hay un df editado guardado y coincide con el filtro actual, usarlo
    # Si no, usar el original
    _cache_key = f"{fecha_desde}_{fecha_hasta}_{filtro_vendedor}"
    if st.session_state.get("_filter_cache_key") != _cache_key:
        # El filtro cambió → resetear el df editado
        st.session_state.df_editado = None
        st.session_state["_filter_cache_key"] = _cache_key

    # df_f es el dataframe activo para los gráficos
    if st.session_state.df_editado is not None:
        df_f = st.session_state.df_editado.copy()
        for col in ["Efectividad_Soles", "Tiempo_PDC"]:
            if col in df_f.columns:
                df_f[col] = pd.to_numeric(df_f[col], errors="coerce").fillna(0)
    else:
        df_f = df_f_original.copy()

    df_f["Fecha_str"] = df_f["Fecha"].dt.date.astype(str)
    ticket_calc = (df_f.groupby(["Vendedor", "Fecha_str"])
        .agg(Ventas_Dia=("Efectividad_Soles", "sum"), Clientes_Dia=("Codigo_PDC", "nunique"))
        .reset_index())
    ticket_calc["Ticket_Calculado"] = ticket_calc.apply(
        lambda r: r["Ventas_Dia"] / r["Clientes_Dia"] if r["Clientes_Dia"] > 0 else 0, axis=1)
    ticket_prom_global = ticket_calc["Ticket_Calculado"].mean() if not ticket_calc.empty else 0

    total_visitas    = len(df_f)
    total_ventas     = df_f["Efectividad_Soles"].sum()
    tiempo_prom      = df_f["Tiempo_PDC"].mean()
    pct_con_terceros = (df_f["Colocacion_Terceros"] == "Sí").mean() * 100 if "Colocacion_Terceros" in df_f.columns else 0
    df_f["Concreto"]  = df_f["Efectividad_Soles"].apply(lambda x: "CONCRETO" if x > 0 else "NO CONCRETO")
    pct_concreto      = (df_f["Concreto"] == "CONCRETO").mean() * 100

    PALETTE_PIE    = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFC107","#00BCD4","#FF69B4"]
    PALETTE_EXHIB  = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFC107","#b0b0b0"]
    PALETTE_GIRO   = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFC107","#00BCD4"]
    PALETTE_CONT   = {"Contaminado": "#e05252", "Limpio": "#4CAF50"}
    FONT_SIZE_AXIS  = 14
    FONT_SIZE_TEXT  = 13
    FONT_SIZE_TITLE = 16

    def base_layout(title=""):
        return dict(
            plot_bgcolor="white", paper_bgcolor="white",
            font=dict(family="DM Sans", size=FONT_SIZE_AXIS),
            margin=dict(t=50, b=40, l=10, r=10),
            title=dict(text=f"<b>{title}</b>", font=dict(size=FONT_SIZE_TITLE, family="DM Sans"), x=0.5, xanchor="center") if title else {},
            xaxis_title="", yaxis_title="",
        )

    st.markdown("---")
    st.markdown("#### 📈 Ticket Promedio por Vendedor y Día")
    st.caption("Ventas totales del vendedor ÷ clientes únicos visitados ese día")
    ticket_display = ticket_calc.rename(columns={
        "Fecha_str": "Fecha", "Ventas_Dia": "Ventas del Día (S/)",
        "Clientes_Dia": "Clientes Visitados", "Ticket_Calculado": "Ticket Promedio (S/)"
    })
    ticket_display["Ventas del Día (S/)"]  = ticket_display["Ventas del Día (S/)"].map("S/ {:,.2f}".format)
    ticket_display["Ticket Promedio (S/)"] = ticket_display["Ticket Promedio (S/)"].map("S/ {:,.2f}".format)
    st.dataframe(ticket_display, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("### 📊 Indicadores Clave")
    k1, k2, k3, k4, k5 = st.columns(5)
    for col, label, value, sub in [
        (k1, "Efectividad Total (S/)",  f"S/ {total_ventas:,.2f}",      "ventas acumuladas"),
        (k2, "Ticket Promedio (S/)",    f"S/ {ticket_prom_global:.2f}", "ventas ÷ clientes / día"),
        (k3, "% Concreto Venta",        f"{pct_concreto:.0f}%",         "visitas con venta"),
        (k4, "Total Visitas",           f"{total_visitas}",             "registros filtrados"),
        (k5, "Tiempo Promedio en PDC",  f"{tiempo_prom:.0f} min",       "promedio por visita"),
    ]:
        with col:
            st.markdown(f"""<div class="kpi-box"><div class="kpi-label">{label}</div>
                <div class="kpi-value">{value}</div><div class="kpi-sub">{sub}</div></div>""",
                unsafe_allow_html=True)

    st.markdown("")

    exhib_cols_dict = {
        "LEGOS_GC": "LEGOS G&C", "TOBOGAN_RITZ_OREO": "TOBOGÁN Ritz/Oreo",
        "EXHIB_KIWI": "EXHIB KIWI", "RITRAZ": "RITRAZ",
        "MEGA_KIWI": "MEGA KIWI", "EXHIBIDOR_OTROS": "OTROS",
    }

    st.markdown("---")
    col_g1, col_g2, col_g3 = st.columns(3)

    with col_g1:
        st.markdown("#### 🗂️ Colocación de Exhibidores")
        data_exhib = []
        for col_name, label in exhib_cols_dict.items():
            if col_name in df_f.columns:
                cnt = int(pd.to_numeric(df_f[col_name], errors="coerce").fillna(0).sum())
                pct = cnt / total_visitas * 100 if total_visitas > 0 else 0
                data_exhib.append({"Exhibidor": label, "Cantidad": cnt, "Pct": round(pct, 1)})
        sin_exhib_mask = df_f[[c for c in exhib_cols_dict if c in df_f.columns]].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) == 0
        sin_cnt = int(sin_exhib_mask.sum())
        sin_pct = sin_cnt / total_visitas * 100 if total_visitas > 0 else 0
        data_exhib.append({"Exhibidor": "SIN EXHIBIDORES", "Cantidad": sin_cnt, "Pct": round(sin_pct, 1)})
        df_exhib = pd.DataFrame(data_exhib)
        df_exhib["Etiqueta"] = df_exhib.apply(lambda r: f"{r['Cantidad']}  ({r['Pct']}%)", axis=1)
        fig_exhib = px.bar(df_exhib, x="Exhibidor", y="Cantidad", text="Etiqueta", color="Exhibidor",
                           color_discrete_sequence=PALETTE_EXHIB)
        fig_exhib.update_traces(textposition="outside", textfont_size=FONT_SIZE_TEXT)
        fig_exhib.update_layout(**base_layout("COLOCACIÓN DE EXHIBIDORES"), showlegend=False,
                                xaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)), yaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)))
        st.plotly_chart(fig_exhib, use_container_width=True)

    with col_g2:
        st.markdown("#### 🏬 Giros de Negocio")
        if "Giro_Negocio" in df_f.columns:
            df_giro = df_f["Giro_Negocio"].value_counts().reset_index()
            df_giro.columns = ["Giro", "Visitas"]
            df_giro["Giro_Short"] = df_giro["Giro"].apply(giro_short)
            total_giro = df_giro["Visitas"].sum()
            df_giro["Label"] = df_giro.apply(
                lambda r: f"{r['Giro_Short'].upper()}\n{r['Visitas']}  ({round(r['Visitas']/total_giro*100)}%)", axis=1)
            fig_giro = go.Figure(go.Pie(
                labels=df_giro["Giro_Short"].str.upper(), values=df_giro["Visitas"],
                text=df_giro["Label"], textinfo="text",
                hovertemplate="<b>%{label}</b><br>Visitas: %{value}<br>%{percent}<extra></extra>",
                hole=0, marker=dict(colors=PALETTE_GIRO[:len(df_giro)], line=dict(color="white", width=2)),
                textfont=dict(size=FONT_SIZE_TEXT, family="DM Sans"), sort=False,
            ))
            fig_giro.update_layout(paper_bgcolor="white", font_family="DM Sans",
                margin=dict(t=50, b=10, l=10, r=10), showlegend=True,
                legend=dict(orientation="v", x=1.02, y=0.5, font=dict(size=FONT_SIZE_AXIS), bgcolor="rgba(0,0,0,0)"),
                title=dict(text="<b>GIROS DE NEGOCIO</b>", font=dict(size=FONT_SIZE_TITLE, family="DM Sans"), x=0.5, xanchor="center"))
            st.plotly_chart(fig_giro, use_container_width=True)

    with col_g3:
        st.markdown("#### 🏷️ Colocación de Terceros")
        if "Colocacion_Terceros" in df_f.columns and "Giro_Negocio" in df_f.columns:
            df_terc_giro = df_f.copy()
            df_terc_giro["Giro_Short"] = df_terc_giro["Giro_Negocio"].apply(giro_short).str.upper()
            df_con = df_terc_giro[df_terc_giro["Colocacion_Terceros"] == "Sí"].groupby("Giro_Short").size().reset_index(name="N")
            sin_n = int((df_terc_giro["Colocacion_Terceros"] != "Sí").sum())
            labels_pie = df_con["Giro_Short"].tolist() + ["SIN COLOCACION"]
            values_pie = df_con["N"].tolist() + [sin_n]
            total_pie  = sum(values_pie)
            text_pie   = [f"{l}\n{v}  ({round(v/total_pie*100)}%)" for l, v in zip(labels_pie, values_pie)]
            colors_pie = PALETTE_PIE[:len(df_con)] + ["#4CAF50"]
            fig_terc = go.Figure(go.Pie(
                labels=labels_pie, values=values_pie, text=text_pie, textinfo="text",
                hovertemplate="<b>%{label}</b><br>Visitas: %{value}<br>%{percent}<extra></extra>",
                hole=0, marker=dict(colors=colors_pie, line=dict(color="white", width=2)),
                textfont=dict(size=FONT_SIZE_TEXT, family="DM Sans"), sort=False,
            ))
            fig_terc.update_layout(paper_bgcolor="white", font_family="DM Sans",
                margin=dict(t=50, b=10, l=10, r=10), showlegend=True,
                legend=dict(orientation="v", x=1.02, y=0.5, font=dict(size=FONT_SIZE_AXIS), bgcolor="rgba(0,0,0,0)"),
                title=dict(text="<b>COLOCACIÓN DE TERCEROS</b>", font=dict(size=FONT_SIZE_TITLE, family="DM Sans"), x=0.5, xanchor="center"))
            st.plotly_chart(fig_terc, use_container_width=True)

    st.markdown("---")

    if "Marca_Tercero" in df_f.columns:
        df_f["Marca_Tercero"] = df_f["Marca_Tercero"].astype(str)
        df_marcas_raw = df_f[
            (df_f["Colocacion_Terceros"] == "Sí") &
            (df_f["Marca_Tercero"].str.strip().str.lower() != "nan") &
            (df_f["Marca_Tercero"].str.strip() != "")
        ]["Marca_Tercero"].copy()
        todas_marcas = []
        for val in df_marcas_raw:
            for m in str(val).split(","):
                m_clean = m.strip().upper()
                if m_clean and m_clean != "NAN":
                    todas_marcas.append(m_clean)

        if todas_marcas:
            from collections import Counter
            conteo = Counter(todas_marcas)
            df_marcas = pd.DataFrame(conteo.items(), columns=["Marca", "Cantidad"]).sort_values("Cantidad", ascending=True)
            total_m = df_marcas["Cantidad"].sum()
            df_marcas["Pct"] = (df_marcas["Cantidad"] / total_m * 100).round(1)
            df_marcas["Etiqueta"] = df_marcas.apply(lambda r: f"  {r['Cantidad']} visitas  ({r['Pct']}%)", axis=1)
            PALETTE_MARCAS = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFC107","#00BCD4","#FF69B4","#8BC34A","#FF5722","#9C27B0","#03A9F4","#CDDC39","#795548","#607D8B"]
            color_marcas = {m: PALETTE_MARCAS[i % len(PALETTE_MARCAS)] for i, m in enumerate(df_marcas["Marca"].tolist())}
            st.markdown("#### 🏷️ Marcas de Terceros Detectadas en PDC")
            st.caption(f"Total de menciones registradas: {total_m}  |  Marcas únicas: {len(df_marcas)}")
            fig_marcas = px.bar(df_marcas, x="Cantidad", y="Marca", orientation="h", text="Etiqueta",
                color="Marca", color_discrete_map=color_marcas,
                title="<b>MARCAS DE COMPETENCIA EN PUNTO DE VENTA</b>")
            fig_marcas.update_traces(textposition="outside", textfont=dict(size=14, family="DM Sans"), marker_line_width=0)
            fig_marcas.update_layout(
                plot_bgcolor="white", paper_bgcolor="white", font=dict(family="DM Sans", size=14),
                title=dict(text="<b>MARCAS DE COMPETENCIA EN PUNTO DE VENTA</b>", font=dict(size=17, family="DM Sans"), x=0.5, xanchor="center"),
                showlegend=False, margin=dict(t=60, b=40, l=20, r=120),
                xaxis=dict(title="Número de Visitas con Presencia", tickfont=dict(size=13), title_font=dict(size=14), gridcolor="#f0f0f0", range=[0, df_marcas["Cantidad"].max() * 1.35]),
                yaxis=dict(title="Marca Competidora", tickfont=dict(size=14), title_font=dict(size=14), automargin=True),
                height=max(300, len(df_marcas) * 55 + 120),
            )
            st.plotly_chart(fig_marcas, use_container_width=True)
        else:
            st.info("No hay marcas de terceros registradas en el período seleccionado.")

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN INTERACTIVA: Tipo de Negocio por Exhibidor
    # Al seleccionar un exhibidor → filtra los gráficos de Contaminación y Visibilidad
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("#### 🏪 Tipo de Negocio por Exhibidor Colocado")
    st.markdown('<div class="filter-info">💡 Selecciona un exhibidor en el selector para filtrar los gráficos de Contaminación y Visibilidad según el tipo de negocio donde fue colocado.</div>', unsafe_allow_html=True)

    # Construir dataframe exhibidor × giro (con nombres cortos)
    exhib_giro_data = []
    for col_ex, label_ex in exhib_cols_dict.items():
        if col_ex in df_f.columns and "Giro_Negocio" in df_f.columns:
            df_ex_sub = df_f[pd.to_numeric(df_f[col_ex], errors="coerce").fillna(0) > 0].copy()
            if df_ex_sub.empty:
                continue
            df_ex_sub["Giro_Short"] = df_ex_sub["Giro_Negocio"].apply(giro_short)
            for giro, grp in df_ex_sub.groupby("Giro_Short"):
                exhib_giro_data.append({
                    "Exhibidor": label_ex,
                    "Giro": giro,
                    "Cantidad": len(grp),
                    "col_key": col_ex,
                })

    if exhib_giro_data:
        df_eg = pd.DataFrame(exhib_giro_data)

        # Totales por exhibidor (suma de todos los giros)
        totales_exhib = df_eg.groupby("Exhibidor")["Cantidad"].sum().reset_index()
        totales_exhib.columns = ["Exhibidor", "Total"]

        giros_unicos = sorted(df_eg["Giro"].unique().tolist())
        color_map_giro = {g: PALETTE_GIRO[i % len(PALETTE_GIRO)] for i, g in enumerate(giros_unicos)}
        df_eg["Etiqueta"] = df_eg["Cantidad"].astype(str)

        # Gráfico apilado con nombres legibles en eje X
        fig_eg = px.bar(
            df_eg, x="Exhibidor", y="Cantidad", color="Giro", text="Etiqueta",
            barmode="stack", color_discrete_map=color_map_giro,
            title="<b>TIPO DE NEGOCIO POR EXHIBIDOR COLOCADO</b>",
        )
        fig_eg.update_traces(textposition="inside", textfont_size=FONT_SIZE_TEXT)
        fig_eg.update_layout(
            **base_layout(),
            title=dict(text="<b>TIPO DE NEGOCIO POR EXHIBIDOR COLOCADO</b>",
                       font=dict(size=FONT_SIZE_TITLE, family="DM Sans"), x=0.5, xanchor="center"),
            xaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS), tickangle=-15),
            yaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS), title="Cantidad"),
            legend_title_text="Giro de Negocio",
            legend=dict(orientation="v", x=1.01, y=1, font=dict(size=12)),
        )
        st.plotly_chart(fig_eg, use_container_width=True)

        # Selector de exhibidor para filtrar los siguientes gráficos
        exhibidores_disponibles = sorted(df_eg["Exhibidor"].unique().tolist())
        opciones_selector = ["Todos (sin filtro)"] + exhibidores_disponibles

        col_sel1, col_sel2 = st.columns([2, 3])
        with col_sel1:
            exhibidor_sel = st.selectbox(
                "🔍 Filtrar Contaminación y Visibilidad por exhibidor:",
                opciones_selector,
                key="sel_exhibidor_filtro",
            )
        with col_sel2:
            if exhibidor_sel != "Todos (sin filtro)":
                # Mostrar distribución de giros para el exhibidor seleccionado
                df_giro_sel = df_eg[df_eg["Exhibidor"] == exhibidor_sel]
                total_sel = df_giro_sel["Cantidad"].sum()
                giros_info = " | ".join([
                    f"{row['Giro']}: {row['Cantidad']} ({round(row['Cantidad']/total_sel*100)}%)"
                    for _, row in df_giro_sel.iterrows()
                ])
                st.markdown(f'<div class="filter-info">📊 <b>{exhibidor_sel}</b> — {total_sel} colocaciones: {giros_info}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="filter-info">Mostrando datos de todos los exhibidores combinados.</div>', unsafe_allow_html=True)

        # Determinar el subconjunto de df_f para filtrar los siguientes gráficos
        if exhibidor_sel != "Todos (sin filtro)":
            # Encontrar la columna correspondiente al exhibidor seleccionado
            col_exhib_sel = None
            for col_key, lbl in exhib_cols_dict.items():
                if lbl == exhibidor_sel:
                    col_exhib_sel = col_key
                    break
            if col_exhib_sel and col_exhib_sel in df_f.columns:
                df_cont_vis = df_f[pd.to_numeric(df_f[col_exhib_sel], errors="coerce").fillna(0) > 0].copy()
                total_cont_vis = len(df_cont_vis)
                st.caption(f"⚠️ Los gráficos de Contaminación y Visibilidad muestran solo las **{total_cont_vis} visitas** donde se colocó **{exhibidor_sel}**.")
            else:
                df_cont_vis = df_f.copy()
                total_cont_vis = total_visitas
        else:
            df_cont_vis = df_f.copy()
            total_cont_vis = total_visitas

    else:
        st.info("No hay datos suficientes para este gráfico.")
        df_cont_vis = df_f.copy()
        total_cont_vis = total_visitas
        exhibidor_sel = "Todos (sin filtro)"

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # CONTAMINACIÓN + VISIBILIDAD — filtrados por exhibidor seleccionado
    # ══════════════════════════════════════════════════════════════════════
    col_cont1, col_cont2 = st.columns(2)

    with col_cont1:
        titulo_cont = "⚠️ Contaminación de Exhibidores"
        if exhibidor_sel != "Todos (sin filtro)":
            titulo_cont += f" — {exhibidor_sel}"
        st.markdown(f"#### {titulo_cont}")
        cont_data = []
        for col_c, label_c in [("CONT_LEGOS_GC", "LEGOS G&C"), ("CONT_TOBOGAN_RITZ_OREO", "TOBOGÁN Ritz/Oreo"), ("CONT_EXHIB_KIWI", "EXHIB KIWI")]:
            if col_c in df_cont_vis.columns:
                serie_c = pd.to_numeric(df_cont_vis[col_c], errors="coerce").fillna(0)
                cont_si  = int(serie_c.sum())
                cont_no  = len(serie_c) - cont_si
                n_total  = len(serie_c)
                pct_si   = cont_si / n_total * 100 if n_total > 0 else 0
                pct_no   = cont_no / n_total * 100 if n_total > 0 else 0
                cont_data.append({"Exhibidor": label_c, "Estado": "Contaminado", "Cantidad": cont_si, "Pct": round(pct_si, 1)})
                cont_data.append({"Exhibidor": label_c, "Estado": "Limpio",      "Cantidad": cont_no, "Pct": round(pct_no, 1)})
        if cont_data:
            df_cont = pd.DataFrame(cont_data)
            df_cont["Etiqueta"] = df_cont.apply(lambda r: f"{r['Cantidad']}  ({r['Pct']}%)" if r["Cantidad"] > 0 else "", axis=1)
            fig_cont = px.bar(df_cont, x="Exhibidor", y="Cantidad", color="Estado", text="Etiqueta",
                barmode="group", color_discrete_map=PALETTE_CONT, category_orders={"Estado": ["Contaminado", "Limpio"]})
            fig_cont.update_traces(textposition="outside", textfont_size=FONT_SIZE_TEXT)
            fig_cont.update_layout(**base_layout("CONTAMINACIÓN DE EXHIBIDORES"),
                xaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)),
                yaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS), title="Cantidad de visitas"),
                legend_title_text="Estado", legend=dict(orientation="h", y=-0.25, font=dict(size=FONT_SIZE_AXIS)))
            st.plotly_chart(fig_cont, use_container_width=True)

    with col_cont2:
        titulo_vis = "👁️ Visibilidad por Exhibidor"
        if exhibidor_sel != "Todos (sin filtro)":
            titulo_vis += f" — {exhibidor_sel}"
        st.markdown(f"#### {titulo_vis}")
        vis_data = []
        for col_v, label_v in [("Visibilidad_Legos","LEGOS G&C"),("Visibilidad_Tobogan","TOBOGÁN Ritz/Oreo"),("Visibilidad_Kiwi","EXHIB KIWI"),("Visibilidad_Otros","OTROS")]:
            if col_v in df_cont_vis.columns:
                serie_v = pd.to_numeric(df_cont_vis[col_v], errors="coerce").fillna(0)
                for nivel, nombre_nivel in [(1, "Alta"), (2, "Media"), (3, "Baja")]:
                    cnt_v = int((serie_v == nivel).sum())
                    pct_v = cnt_v / total_cont_vis * 100 if total_cont_vis > 0 else 0
                    vis_data.append({"Exhibidor": label_v, "Nivel": nombre_nivel, "Cantidad": cnt_v, "Pct": round(pct_v, 1)})
        if vis_data:
            df_vis = pd.DataFrame(vis_data)
            df_vis["Etiqueta"] = df_vis.apply(lambda r: f"{r['Cantidad']}  ({r['Pct']}%)" if r["Cantidad"] > 0 else "", axis=1)
            fig_vis = px.bar(df_vis, x="Exhibidor", y="Cantidad", color="Nivel", text="Etiqueta",
                barmode="group", color_discrete_map={"Alta": "#4CAF50", "Media": "#FFC107", "Baja": "#e05252"},
                category_orders={"Nivel": ["Alta", "Media", "Baja"]})
            fig_vis.update_traces(textposition="outside", textfont_size=FONT_SIZE_TEXT)
            fig_vis.update_layout(**base_layout("VISIBILIDAD POR EXHIBIDOR"),
                xaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)),
                yaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS), title="Cantidad de visitas"),
                legend_title_text="Visibilidad", legend=dict(orientation="h", y=-0.2, font=dict(size=FONT_SIZE_AXIS)))
            st.plotly_chart(fig_vis, use_container_width=True)

    st.markdown("---")
    col_g4, col_g5 = st.columns(2)

    with col_g4:
        st.markdown("#### 📊 Efectividad de Visitas")
        df_efec = df_f["Concreto"].value_counts().reset_index()
        df_efec.columns = ["Estado", "Cantidad"]
        total_efec = df_efec["Cantidad"].sum()
        df_efec["Etiqueta"] = df_efec.apply(lambda r: f"{round(r['Cantidad']/total_efec*100)}%  ({r['Cantidad']})", axis=1)
        fig_efec = px.bar(df_efec, y="Estado", x="Cantidad", orientation="h", text="Etiqueta",
                          color="Estado", color_discrete_map={"CONCRETO": "#7b5ea7", "NO CONCRETO": "#e05252"})
        fig_efec.update_traces(textposition="outside", textfont_size=FONT_SIZE_TEXT)
        fig_efec.update_layout(**base_layout("EFECTIVIDAD DE VISITAS"), showlegend=False,
            xaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)), yaxis=dict(tickfont=dict(size=FONT_SIZE_AXIS)))
        st.plotly_chart(fig_efec, use_container_width=True)

    with col_g5:
        st.markdown("#### 📍 Mapa de Visitas")
        if "Latitud" in df_f.columns and "Longitud" in df_f.columns:
            df_map = df_f.copy()
            df_map["Latitud"]  = pd.to_numeric(df_map["Latitud"],  errors="coerce")
            df_map["Longitud"] = pd.to_numeric(df_map["Longitud"], errors="coerce")
            df_map = df_map.dropna(subset=["Latitud","Longitud"])
            if not df_map.empty:
                zonas_unicas = df_map["Zona"].dropna().unique().tolist() if "Zona" in df_map.columns else []
                palette_map  = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFBF00","#2196F3","#00BCD4"]
                zona_color   = {z: palette_map[i % len(palette_map)] for i, z in enumerate(zonas_unicas)}
                fig_map = go.Figure()
                for zona, grp in df_map.groupby("Zona"):
                    ch = zona_color.get(zona, "#888888")
                    r2,g2,b2 = int(ch[1:3],16), int(ch[3:5],16), int(ch[5:7],16)
                    fig_map.add_trace(go.Scattermapbox(lat=grp["Latitud"].tolist(), lon=grp["Longitud"].tolist(),
                        mode="markers", marker=dict(size=40, color=f"rgba({r2},{g2},{b2},0.18)"),
                        hoverinfo="skip", showlegend=False, name=f"sombra{zona}"))
                for zona, grp in df_map.groupby("Zona"):
                    ch = zona_color.get(zona, "#888888")
                    giro_col = grp["Giro_Negocio"].apply(giro_short) if "Giro_Negocio" in grp.columns else grp.index.astype(str)
                    fig_map.add_trace(go.Scattermapbox(lat=grp["Latitud"].tolist(), lon=grp["Longitud"].tolist(),
                        mode="markers", marker=dict(size=13, color=ch), name=zona,
                        text=grp["Nombre_Cliente"].tolist(),
                        customdata=list(zip(grp["Zona"].tolist(), grp["Efectividad_Soles"].tolist(), giro_col.tolist())),
                        hovertemplate="<b>%{text}</b><br>Zona: %{customdata[0]}<br>Venta: S/ %{customdata[1]}<br>Giro: %{customdata[2]}<extra></extra>"))
                lat_c   = df_map["Latitud"].mean()
                lon_c   = df_map["Longitud"].mean()
                lat_rng = df_map["Latitud"].max() - df_map["Latitud"].min()
                zoom_auto = 13 if lat_rng < 0.01 else (11 if lat_rng < 0.05 else 9)
                fig_map.update_layout(
                    mapbox=dict(style="open-street-map", center=dict(lat=lat_c, lon=lon_c), zoom=zoom_auto),
                    margin=dict(t=0,b=0,l=0,r=0), height=380,
                    legend=dict(title="Zona", bgcolor="rgba(255,255,255,0.85)", bordercolor="#ddd", borderwidth=1),
                    font_family="DM Sans")
                st.plotly_chart(fig_map, use_container_width=True)
            else:
                st.info("Agrega visitas con coordenadas para ver el mapa.")

    st.markdown("---")
    st.markdown("#### 📦 Presencia de Productos (% de visitas)")

    def render_presencia_bar(productos_dict, df_source, total_vis, palette):
        presencia_pct = []
        for key, label in productos_dict.items():
            if key in df_source.columns:
                serie     = pd.to_numeric(df_source[key], errors="coerce")
                total_con = int(serie.sum())
                pct       = serie.mean() * 100
                presencia_pct.append({"Producto": label, "Presencia %": round(pct, 1),
                    "Etiqueta": f"{round(pct,1)}%  ({total_con}/{total_vis})"})
        if not presencia_pct:
            st.info("Sin datos.")
            return
        df_pres = pd.DataFrame(presencia_pct).sort_values("Presencia %", ascending=False)
        color_map_p = {row["Producto"]: palette[i % len(palette)] for i, (_, row) in enumerate(df_pres.iterrows())}
        fig_pres = px.bar(df_pres, x="Producto", y="Presencia %", color="Producto",
            color_discrete_map=color_map_p, range_y=[0, 115], text="Etiqueta")
        fig_pres.update_traces(textposition="outside", textfont_size=FONT_SIZE_TEXT)
        fig_pres.update_layout(**base_layout(), showlegend=False,
            xaxis=dict(title="", tickangle=-30, tickfont=dict(size=FONT_SIZE_AXIS)),
            yaxis=dict(title="Presencia %", range=[0, 115], tickfont=dict(size=FONT_SIZE_AXIS)))
        st.plotly_chart(fig_pres, use_container_width=True)

    PALETTE_PRODUCTOS = ["#4472C4","#e05252","#6a9e4f","#FF7F0E","#7b5ea7","#FFC107","#00BCD4","#FF69B4","#8BC34A","#FF5722"]

    tab_biscuits, tab_foco, tab_gyc = st.tabs(["🍪 Biscuits", "⭐ Productos Foco", "🍬 G&C"])
    with tab_biscuits:
        render_presencia_bar({"OREO_34GR":"OREO 34GR","OREO_54GR":"OREO 54GR","OREO_ROLLO":"OREO ROLLO","RITZ_ROLLO":"RITZ ROLLO","RITZ_PACK":"RITZ PACK","FIELD_CC":"FIELD CC","FIELD_DP":"FIELD DP","FIELD_VAIN":"FIELD VAIN","CLUB_SOCIAL_TRA":"CLUB SOCIAL TRA"}, df_f, total_visitas, PALETTE_PRODUCTOS)
    with tab_foco:
        render_presencia_bar({"OREO_FRESA_PACK":"OREO FRESA (Pack)","OREO_FRESA_ROLLO":"OREO FRESA (Rollo)","OREO_CHOCO_LIMON_PACK":"OREO CHOCO LIMÓN (Pack)","OREO_CHOCO_LIMON_ROLLO":"OREO CHOCO LIMÓN (Rollo)","CLUB_SOCIAL_SAB":"CLUB SOCIAL (Sabores)","ROLLO_GOLDEN":"OREO GOLDEN (Rollo)","ROLLO_CHOCOLATE":"OREO CHOCOLATE (Rollo)"}, df_f, total_visitas, PALETTE_PRODUCTOS)
    with tab_gyc:
        render_presencia_bar({"TRIDENT_5s":"TRIDENT 5s","TRIDENT_EVUP":"TRIDENT EVUP","HALLS_12s":"HALLS 12s","HALLS_100s":"HALLS 100s","CHICLETS_2S":"CHICLETS 2S","BUBBALOO":"BUBBALOO"}, df_f, total_visitas, PALETTE_PRODUCTOS)

    # ══════════════════════════════════════════════════════════════════════
    # TABLA EDITABLE — Los cambios actualizan los gráficos al guardar
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 📋 Tabla de Visitas — Editable")
    st.markdown('<div class="edit-info">✏️ <b>Puedes editar los datos directamente en la tabla.</b> Haz doble clic en cualquier celda para modificarla. Cuando termines, presiona <b>"✅ Aplicar cambios y actualizar gráficos"</b> para que los gráficos reflejen las ediciones. Los cambios también se guardan en el archivo CSV.</div>', unsafe_allow_html=True)

    # Columnas editables que se muestran en la tabla
    cols_editables = [
        "Fecha", "Codigo_PDC", "Nombre_Cliente", "Giro_Negocio",
        "Vendedor", "Codigo_Vendedor", "Mesa", "Zona",
        "Efectividad_Soles", "Tiempo_PDC",
        "Visibilidad_Legos", "Visibilidad_Tobogan", "Visibilidad_Kiwi", "Visibilidad_Otros",
        "Colocacion_Terceros", "Marca_Tercero",
        "LEGOS_GC", "TOBOGAN_RITZ_OREO", "EXHIB_KIWI", "RITRAZ", "MEGA_KIWI", "EXHIBIDOR_OTROS",
        "CONT_LEGOS_GC", "CONT_TOBOGAN_RITZ_OREO", "CONT_EXHIB_KIWI",
    ]
    cols_mostrar = [c for c in cols_editables if c in df_f.columns]

    # Preparar datos para el editor
    df_para_editar = df_f[cols_mostrar].copy()
    df_para_editar["Fecha"] = df_para_editar["Fecha"].dt.strftime("%Y-%m-%d") if pd.api.types.is_datetime64_any_dtype(df_para_editar["Fecha"]) else df_para_editar["Fecha"].astype(str)

    # Configuración de columnas para el editor
    column_config = {
        "Fecha": st.column_config.TextColumn("Fecha", help="Formato YYYY-MM-DD"),
        "Codigo_PDC": st.column_config.TextColumn("Código PDC"),
        "Nombre_Cliente": st.column_config.TextColumn("Cliente"),
        "Giro_Negocio": st.column_config.SelectboxColumn(
            "Giro de Negocio",
            options=[
                "1 - Bodega",
                "2 - Minimarket / Tiendas",
                "3 - Kiosko",
                "4 - Especializados (Panificadora, Horeca, Internet...)",
                "5 - Otros (Puesto de mercado, Centros Educativos...)",
            ],
        ),
        "Vendedor": st.column_config.TextColumn("Vendedor"),
        "Codigo_Vendedor": st.column_config.TextColumn("Cód. Vendedor"),
        "Mesa": st.column_config.TextColumn("Mesa"),
        "Zona": st.column_config.TextColumn("Zona"),
        "Efectividad_Soles": st.column_config.NumberColumn("Efectividad (S/)", min_value=0.0, format="%.2f"),
        "Tiempo_PDC": st.column_config.NumberColumn("Tiempo PDC (min)", min_value=0, step=1),
        "Visibilidad_Legos":   st.column_config.SelectboxColumn("Vis. LEGOS",   options=[1, 2, 3]),
        "Visibilidad_Tobogan": st.column_config.SelectboxColumn("Vis. TOBOGÁN", options=[1, 2, 3]),
        "Visibilidad_Kiwi":    st.column_config.SelectboxColumn("Vis. KIWI",    options=[1, 2, 3]),
        "Visibilidad_Otros":   st.column_config.SelectboxColumn("Vis. OTROS",   options=[1, 2, 3]),
        "Colocacion_Terceros": st.column_config.SelectboxColumn("¿Terceros?", options=["No", "Sí"]),
        "Marca_Tercero": st.column_config.TextColumn("Marcas Terceros"),
        "LEGOS_GC":          st.column_config.CheckboxColumn("LEGOS G&C"),
        "TOBOGAN_RITZ_OREO": st.column_config.CheckboxColumn("TOBOGÁN"),
        "EXHIB_KIWI":        st.column_config.CheckboxColumn("EXHIB KIWI"),
        "RITRAZ":            st.column_config.CheckboxColumn("RITRAZ"),
        "MEGA_KIWI":         st.column_config.CheckboxColumn("MEGA KIWI"),
        "EXHIBIDOR_OTROS":   st.column_config.CheckboxColumn("EXH. OTROS"),
        "CONT_LEGOS_GC":          st.column_config.CheckboxColumn("Cont. LEGOS"),
        "CONT_TOBOGAN_RITZ_OREO": st.column_config.CheckboxColumn("Cont. TOBOGÁN"),
        "CONT_EXHIB_KIWI":        st.column_config.CheckboxColumn("Cont. KIWI"),
    }

    # Para columnas booleanas (checkboxes), convertir 0/1 a bool
    bool_cols = ["LEGOS_GC","TOBOGAN_RITZ_OREO","EXHIB_KIWI","RITRAZ","MEGA_KIWI","EXHIBIDOR_OTROS",
                 "CONT_LEGOS_GC","CONT_TOBOGAN_RITZ_OREO","CONT_EXHIB_KIWI"]
    for bc in bool_cols:
        if bc in df_para_editar.columns:
            df_para_editar[bc] = pd.to_numeric(df_para_editar[bc], errors="coerce").fillna(0).astype(bool)

    edited_df = st.data_editor(
        df_para_editar,
        column_config={k: v for k, v in column_config.items() if k in df_para_editar.columns},
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="tabla_editable_principal",
    )

    col_apply1, col_apply2, col_apply3 = st.columns([2, 2, 3])
    with col_apply1:
        if st.button("✅ Aplicar cambios y actualizar gráficos", use_container_width=True, key="btn_apply_edit"):
            # Convertir bools de vuelta a int
            df_aplicado = edited_df.copy()
            for bc in bool_cols:
                if bc in df_aplicado.columns:
                    df_aplicado[bc] = df_aplicado[bc].astype(int)
            # Reconstruir fechas
            df_aplicado["Fecha"] = pd.to_datetime(df_aplicado["Fecha"], errors="coerce")

            # Actualizar el CSV: reemplazar las filas del rango filtrado
            df_completo = cargar_datos()
            df_completo["Fecha"] = pd.to_datetime(df_completo["Fecha"])

            # Identificar índices originales del filtro
            mask_orig = (df_completo["Fecha"].dt.date >= fecha_desde) & (df_completo["Fecha"].dt.date <= fecha_hasta)
            if filtro_vendedor != "Todos":
                mask_orig = mask_orig & (df_completo["Vendedor"] == filtro_vendedor)

            # Alinear las columnas del df_aplicado con el df_completo
            df_aplicado_full = df_completo[mask_orig].copy()
            for col_ed in df_aplicado.columns:
                if col_ed in df_aplicado_full.columns:
                    df_aplicado_full[col_ed] = df_aplicado[col_ed].values[:len(df_aplicado_full)]

            df_completo.loc[mask_orig, df_aplicado.columns.tolist()] = df_aplicado_full[df_aplicado.columns.tolist()].values
            df_completo.to_csv(CSV_FILE, index=False)

            # Guardar el df editado en session_state para que los gráficos lo usen
            df_para_graficos = df_aplicado.copy()
            df_para_graficos["Fecha"] = pd.to_datetime(df_para_graficos["Fecha"], errors="coerce")
            for col_num in ["Efectividad_Soles", "Tiempo_PDC"]:
                if col_num in df_para_graficos.columns:
                    df_para_graficos[col_num] = pd.to_numeric(df_para_graficos[col_num], errors="coerce").fillna(0)
            # Añadir columnas que no están en la tabla editable pero sí en df_f
            for col_falt in df_f.columns:
                if col_falt not in df_para_graficos.columns:
                    df_para_graficos[col_falt] = df_f[col_falt].values[:len(df_para_graficos)]
            st.session_state.df_editado = df_para_graficos
            st.success("✅ Cambios aplicados. Los gráficos se actualizarán al recargar.")
            st.rerun()

    with col_apply2:
        if st.session_state.df_editado is not None:
            if st.button("↩️ Restaurar datos originales", use_container_width=True, key="btn_restore"):
                st.session_state.df_editado = None
                st.rerun()

    with col_apply3:
        if st.session_state.df_editado is not None:
            st.markdown('<span style="color:#6a9e4f;font-size:13px;">✔ Mostrando datos editados</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span style="color:#aaa;font-size:13px;">Mostrando datos originales del CSV</span>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### ⬇️ Descargas")

    dcol1, dcol2, dcol3 = st.columns(3)

    with dcol1:
        buffer = io.BytesIO()
        export_df = df_f.copy()
        export_df = export_df.merge(ticket_calc[["Vendedor","Fecha_str","Ticket_Calculado"]], on=["Vendedor","Fecha_str"], how="left")
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            export_df.drop(columns=["Imagen_Path","Fecha_str","Concreto"], errors="ignore").to_excel(writer, index=False, sheet_name="Visitas")
        buffer.seek(0)
        st.download_button(label="⬇️ Descargar Excel de Datos", data=buffer,
            file_name=f"visitas_MDZ_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    with dcol2:
        try:
            with st.spinner("Generando Dashboard con gráficos..."):
                dash_buf = generar_dashboard_excel(
                    df_f, ticket_calc, fecha_desde, fecha_hasta, filtro_vendedor,
                    total_visitas, total_ventas, ticket_prom_global,
                    tiempo_prom, pct_con_terceros, pct_concreto, data_exhib
                )
            st.download_button(
                label="📊 Descargar Dashboard Operativo (Excel)",
                data=dash_buf,
                file_name=f"dashboard_operativo_MDZ_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error generando dashboard: {e}")
            import traceback
            st.code(traceback.format_exc())

    with dcol3:
        imagenes = []
        if "Imagen_Path" in df_f.columns:
            for p_raw in df_f["Imagen_Path"].dropna().tolist():
                if p_raw:
                    for p in str(p_raw).split("|"):
                        if p.strip() and os.path.exists(p.strip()):
                            imagenes.append(p.strip())
        if imagenes:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for img_path in imagenes:
                    zf.write(img_path, os.path.basename(img_path))
            zip_buffer.seek(0)
            st.download_button(label=f"🖼️ Descargar imágenes ({len(imagenes)} fotos)", data=zip_buffer,
                file_name=f"imagenes_MDZ_{date.today()}.zip", mime="application/zip", use_container_width=True)
        else:
            st.info("No hay imágenes guardadas aún.")
